#!/usr/bin/env python3
"""Analyze Excel sheets for potential merging."""

from openpyxl import load_workbook
from collections import defaultdict
from urllib.parse import urlparse

EXCEL_FILE = "D:/KB-OPEN/github-repos/SaveToWaybackMachine/archived-sites/mmdc.nl/mmdc-urls-spider-output.xlsx"

wb = load_workbook(EXCEL_FILE, read_only=True)

print("=" * 70)
print("SHEET ANALYSIS - Potential Merges")
print("=" * 70)

sheet_data = {}

for sheet in wb.sheetnames:
    if sheet == "SUMMARY":
        continue
    ws = wb[sheet]
    urls = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            urls.append(str(row[0]))

    sheet_data[sheet] = urls

    if not urls:
        print(f"\n{sheet}: 0 URLs")
        continue

    # Analyze URL patterns
    patterns = defaultdict(int)
    base_paths = set()
    for url in urls:
        path = urlparse(url).path
        parts = [p for p in path.split('/') if p]
        if len(parts) >= 2:
            base_paths.add(f"/{parts[0]}/{parts[1]}/")
        if len(parts) >= 3:
            pattern = f"/{parts[0]}/{parts[1]}/{parts[2][:20]}..."
        elif len(parts) >= 2:
            pattern = f"/{parts[0]}/{parts[1]}/"
        else:
            pattern = path[:50]
        patterns[pattern] += 1

    print(f"\n{sheet}: {len(urls)} URLs")
    print(f"  Base paths: {', '.join(sorted(base_paths)[:3])}")
    for pattern, count in sorted(patterns.items(), key=lambda x: -x[1])[:3]:
        print(f"    {pattern}: {count}")

wb.close()

print("\n" + "=" * 70)
print("MERGE RECOMMENDATIONS")
print("=" * 70)

print("""
Based on the URL structure analysis:

1. KEEP SEPARATE (distinct site sections):
   - ENTRY_POINT (homepage)
   - SEARCH_CATALOG (search UI)
   - CATALOG_RECORDS (11,738 manuscript records - main content!)
   - PDF_DOCUMENTS (downloadable files)
   - STATIC_ASSETS (CSS, JS, images)

2. COULD MERGE → "STATIC_CONTENT" or "SITE_PAGES":
   - HIGHLIGHTS (/static/site/highlights/)
   - RESEARCH_EDUCATION (/static/site/research_and_education/)
   - LITERATURE (/static/site/literature/)
   - COLLECTIONS (/static/site/collections/)
   - LINKS (/static/site/links/)
   - ABOUT (/static/site/about/)
   - OTHER (misc pages)

   Reason: All are static HTML content pages under /static/site/

3. MANUSCRIPT_RECORDS can be REMOVED (0 URLs - redundant with CATALOG_RECORDS)

SUGGESTED NEW STRUCTURE:
   - SUMMARY (overview)
   - CATALOG_RECORDS (11,738) - The main manuscript database
   - SITE_PAGES (~300) - All static content merged
   - PDF_DOCUMENTS (63) - Downloadable files
   - STATIC_ASSETS (101) - Technical assets
""")
