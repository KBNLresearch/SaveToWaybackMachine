#!/usr/bin/env python3
"""
Create a unified Excel file with all URLs in a single sheet.
Adds a 'Section' column to indicate the origin of each URL.
"""

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse, unquote

BASE_DIR = Path(__file__).parent.parent
OLD_EXCEL = BASE_DIR / "mmdc-urls-spider-output.xlsx"
NEW_EXCEL = BASE_DIR / "mmdc-urls-UNIFIED.xlsx"
CATALOG_IDS = BASE_DIR / "catalog-record-ids.txt"
PDF_URLS = BASE_DIR / "pdf-urls.txt"

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CENTER = Alignment(horizontal="center")
WRAP = Alignment(wrap_text=True)


def classify_url(url: str) -> str:
    """Classify URL into section based on path."""
    path = urlparse(url).path.lower()

    # Check for search/catalog detail pages
    if "search/detail" in path or "recordId=" in url:
        return "CATALOG_RECORDS"

    # Static assets
    if any(ext in path for ext in ['.css', '.js', '.png', '.jpg', '.jpeg', '.gif', '.svg', '.ico', '.woff', '.ttf']):
        return "STATIC_ASSETS"

    # PDFs
    if '.pdf' in path.lower():
        return "PDF_DOCUMENTS"

    # Site sections
    if '/highlights/' in path:
        return "HIGHLIGHTS"
    if '/research_and_education/' in path or '/research_education/' in path:
        return "RESEARCH_EDUCATION"
    if '/literature/' in path:
        return "LITERATURE"
    if '/collections/' in path:
        return "COLLECTIONS"
    if '/links/' in path:
        return "LINKS"
    if '/about/' in path:
        return "ABOUT"
    if '/search/' in path:
        return "SEARCH_UI"

    # Homepage
    if path in ['/', '', '/index.html', '/static/site/index.html']:
        return "HOMEPAGE"

    return "OTHER"


def get_url_title(url: str) -> str:
    """Extract a readable title from URL path."""
    path = urlparse(url).path
    path = unquote(path)  # Decode %20 etc.

    # Get last meaningful segment
    parts = [p for p in path.split('/') if p and p != 'index.html']
    if parts:
        title = parts[-1]
        # Clean up
        title = title.replace('.html', '').replace('_', ' ').replace('-', ' ')
        return title.title()
    return "Homepage"


def main():
    print("=" * 70)
    print("Creating Unified Excel with Single Sheet")
    print("=" * 70)

    all_urls = []

    # 1. Load URLs from existing Excel (static pages)
    print("\n1. Loading static pages from existing Excel...")
    if OLD_EXCEL.exists():
        wb_old = load_workbook(OLD_EXCEL, read_only=True)
        for sheet in wb_old.sheetnames:
            if sheet in ["SUMMARY", "CATALOG_RECORDS", "PDF_DOCUMENTS"]:
                continue  # Skip these, we'll add them separately
            ws = wb_old[sheet]
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0]:
                    url = str(row[0])
                    section = classify_url(url)
                    all_urls.append({
                        "url": url,
                        "section": section,
                        "title": get_url_title(url),
                    })
        wb_old.close()
        print(f"   Loaded {len(all_urls)} static page URLs")

    # 2. Load catalog record URLs
    print("\n2. Loading catalog records from IDs file...")
    if CATALOG_IDS.exists():
        ids = CATALOG_IDS.read_text().strip().split("\n")
        url_template = "https://mmdc.nl/static/site/search/detail.html?recordId={id}#r{id}"
        catalog_count = 0
        for record_id in ids:
            url = url_template.format(id=record_id)
            all_urls.append({
                "url": url,
                "section": "CATALOG_RECORDS",
                "title": f"Record {record_id}",
            })
            catalog_count += 1
        print(f"   Loaded {catalog_count} catalog record URLs")

    # 3. Load PDF URLs
    print("\n3. Loading PDF URLs...")
    if PDF_URLS.exists():
        pdfs = PDF_URLS.read_text().strip().split("\n")
        pdf_count = 0
        for url in pdfs:
            if url:
                all_urls.append({
                    "url": url,
                    "section": "PDF_DOCUMENTS",
                    "title": get_url_title(url),
                })
                pdf_count += 1
        print(f"   Loaded {pdf_count} PDF URLs")

    # Remove duplicates (keep first occurrence)
    seen = set()
    unique_urls = []
    for item in all_urls:
        if item["url"] not in seen:
            seen.add(item["url"])
            unique_urls.append(item)

    print(f"\n   Total unique URLs: {len(unique_urls)}")

    # 4. Create new Excel
    print("\n4. Creating unified Excel...")
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create ALL_URLS sheet
    ws = wb.create_sheet("ALL_URLS")

    # Headers
    headers = ["URL", "Section", "Title", "Status", "Notes"]
    ws.append(headers)

    # Add data rows
    for item in unique_urls:
        ws.append([
            item["url"],
            item["section"],
            item["title"],
            "pending",
            ""
        ])

    # Column widths
    ws.column_dimensions['A'].width = 90  # URL
    ws.column_dimensions['B'].width = 18  # Section
    ws.column_dimensions['C'].width = 40  # Title
    ws.column_dimensions['D'].width = 10  # Status
    ws.column_dimensions['E'].width = 30  # Notes

    # 5. Create SUMMARY sheet
    ws_summary = wb.create_sheet("SUMMARY", 0)

    # Count by section
    section_counts = {}
    for item in unique_urls:
        section = item["section"]
        section_counts[section] = section_counts.get(section, 0) + 1

    ws_summary['A1'] = "MMDC.nl URL Inventory - Unified Export"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:C1')

    ws_summary['A3'] = "Generated:"
    ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary['A4'] = "Site shutdown:"
    ws_summary['B4'] = "December 15, 2025"
    ws_summary['B4'].font = Font(bold=True, color="FF0000")

    ws_summary['A6'] = "Section"
    ws_summary['B6'] = "Count"
    ws_summary['A6'].font = HEADER_FONT
    ws_summary['A6'].fill = HEADER_FILL
    ws_summary['B6'].font = HEADER_FONT
    ws_summary['B6'].fill = HEADER_FILL

    row = 7
    for section, count in sorted(section_counts.items(), key=lambda x: -x[1]):
        ws_summary[f'A{row}'] = section
        ws_summary[f'B{row}'] = count
        ws_summary[f'B{row}'].alignment = CENTER
        row += 1

    ws_summary[f'A{row}'] = "TOTAL"
    ws_summary[f'B{row}'] = len(unique_urls)
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'].alignment = CENTER

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 12

    # Save
    wb.save(NEW_EXCEL)
    wb.close()

    print(f"\n   Saved: {NEW_EXCEL}")

    # 6. Also export to CSV
    csv_file = BASE_DIR / "mmdc-urls-ALL.csv"
    print(f"\n5. Exporting to CSV: {csv_file}")

    import csv
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["URL", "Section", "Title", "Status", "Notes"])
        for item in unique_urls:
            writer.writerow([item["url"], item["section"], item["title"], "pending", ""])

    print(f"   Saved: {csv_file}")

    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    for section, count in sorted(section_counts.items(), key=lambda x: -x[1]):
        print(f"  {section}: {count:,}")
    print(f"  {'='*30}")
    print(f"  TOTAL: {len(unique_urls):,}")
    print("=" * 70)


if __name__ == "__main__":
    main()
