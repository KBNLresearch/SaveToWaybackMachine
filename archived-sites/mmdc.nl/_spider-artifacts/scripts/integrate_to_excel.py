#!/usr/bin/env python3
"""
Integrate all URL sources into the main Excel file.

Adds new sheets:
- CATALOG_RECORDS: 11,738 catalog detail pages from SRU API
- PDF_DOCUMENTS: 63 PDF files
- SUMMARY: Overview of all URL categories
"""

from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
EXCEL_FILE = BASE_DIR / "mmdc-urls-spider-output.xlsx"
CATALOG_IDS_FILE = BASE_DIR / "catalog-record-ids.txt"
PDF_URLS_FILE = BASE_DIR / "pdf-urls.txt"

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
CENTER = Alignment(horizontal="center")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def add_catalog_records(wb):
    """Add CATALOG_RECORDS sheet with all catalog URLs."""
    print("Adding CATALOG_RECORDS sheet...")

    # Remove if exists
    if "CATALOG_RECORDS" in wb.sheetnames:
        del wb["CATALOG_RECORDS"]

    ws = wb.create_sheet("CATALOG_RECORDS")

    # Headers
    headers = ["Record ID", "URL", "Source", "Added"]
    ws.append(headers)

    for col, cell in enumerate(ws[1], 1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # Load record IDs
    ids = CATALOG_IDS_FILE.read_text().strip().split("\n")
    timestamp = datetime.now().isoformat()

    # Add rows
    url_template = "https://mmdc.nl/static/site/search/detail.html?recordId={id}#r{id}"
    for record_id in ids:
        ws.append([
            int(record_id),
            url_template.format(id=record_id),
            "SRU API (jsru.kb.nl)",
            timestamp
        ])

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 22

    print(f"  Added {len(ids):,} catalog records")
    return len(ids)


def add_pdf_documents(wb):
    """Add PDF_DOCUMENTS sheet."""
    print("Adding PDF_DOCUMENTS sheet...")

    if "PDF_DOCUMENTS" in wb.sheetnames:
        del wb["PDF_DOCUMENTS"]

    ws = wb.create_sheet("PDF_DOCUMENTS")

    # Headers
    headers = ["URL", "Filename", "Source", "Added"]
    ws.append(headers)

    for col, cell in enumerate(ws[1], 1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # Load PDFs
    if PDF_URLS_FILE.exists():
        pdfs = PDF_URLS_FILE.read_text().strip().split("\n")
        pdfs = [p for p in pdfs if p]
    else:
        pdfs = []

    timestamp = datetime.now().isoformat()

    for url in pdfs:
        filename = url.split("/")[-1]
        ws.append([url, filename, "Page crawl extraction", timestamp])

    # Column widths
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 22

    print(f"  Added {len(pdfs):,} PDF documents")
    return len(pdfs)


def add_summary_sheet(wb, counts):
    """Add SUMMARY sheet with overview."""
    print("Adding SUMMARY sheet...")

    if "SUMMARY" in wb.sheetnames:
        del wb["SUMMARY"]

    # Insert at beginning
    ws = wb.create_sheet("SUMMARY", 0)

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "MMDC.nl URL Inventory - Complete Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = CENTER

    # Generation info
    ws['A3'] = "Generated:"
    ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A4'] = "Site shutdown:"
    ws['B4'] = "December 15, 2025"
    ws['B4'].font = Font(bold=True, color="FF0000")

    # URL counts table
    ws['A6'] = "URL Category"
    ws['B6'] = "Sheet"
    ws['C6'] = "Count"

    for col in ['A', 'B', 'C']:
        ws[f'{col}6'].font = HEADER_FONT
        ws[f'{col}6'].fill = HEADER_FILL
        ws[f'{col}6'].alignment = CENTER

    # Data rows
    row = 7
    total = 0

    categories = [
        ("Entry Point", "ENTRY_POINT", counts.get("ENTRY_POINT", 0)),
        ("Search/Catalog UI", "SEARCH_CATALOG", counts.get("SEARCH_CATALOG", 0)),
        ("Highlights", "HIGHLIGHTS", counts.get("HIGHLIGHTS", 0)),
        ("Research & Education", "RESEARCH_EDUCATION", counts.get("RESEARCH_EDUCATION", 0)),
        ("Literature", "LITERATURE", counts.get("LITERATURE", 0)),
        ("Collections", "COLLECTIONS", counts.get("COLLECTIONS", 0)),
        ("Links", "LINKS", counts.get("LINKS", 0)),
        ("About", "ABOUT", counts.get("ABOUT", 0)),
        ("Manuscript Records (Static)", "MANUSCRIPT_RECORDS", counts.get("MANUSCRIPT_RECORDS", 0)),
        ("Static Assets", "STATIC_ASSETS", counts.get("STATIC_ASSETS", 0)),
        ("Other Pages", "OTHER", counts.get("OTHER", 0)),
        ("Catalog Records (SRU API)", "CATALOG_RECORDS", counts.get("CATALOG_RECORDS", 0)),
        ("PDF Documents", "PDF_DOCUMENTS", counts.get("PDF_DOCUMENTS", 0)),
    ]

    for name, sheet, count in categories:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = sheet
        ws[f'C{row}'] = count
        ws[f'C{row}'].alignment = CENTER
        if count > 1000:
            ws[f'A{row}'].fill = SUMMARY_FILL
            ws[f'B{row}'].fill = SUMMARY_FILL
            ws[f'C{row}'].fill = SUMMARY_FILL
        total += count
        row += 1

    # Total row
    ws[f'A{row}'] = "TOTAL URLS"
    ws[f'B{row}'] = ""
    ws[f'C{row}'] = total
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = TOTAL_FILL
    ws[f'C{row}'].alignment = CENTER

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12

    # Key discovery note
    row += 2
    ws[f'A{row}'] = "Key Discovery:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "11,738 catalog records extracted via KB's SRU API (jsru.kb.nl)"
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Query: dcterms.isPartOf adj PTP"

    print(f"  Total URLs: {total:,}")
    return total


def main():
    """Integrate all sources into Excel."""
    print("=" * 60)
    print("Integrating All URLs into Excel")
    print("=" * 60)

    # Load workbook
    wb = load_workbook(EXCEL_FILE)

    # Get existing sheet counts
    counts = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        count = ws.max_row - 1 if ws.max_row > 1 else 0
        counts[sheet_name] = count
        print(f"  Existing: {sheet_name} = {count}")

    # Add new sheets
    counts["CATALOG_RECORDS"] = add_catalog_records(wb)
    counts["PDF_DOCUMENTS"] = add_pdf_documents(wb)

    # Add summary
    total = add_summary_sheet(wb, counts)

    # Save
    wb.save(EXCEL_FILE)
    wb.close()

    print("=" * 60)
    print(f"Excel updated: {EXCEL_FILE}")
    print(f"Total URLs: {total:,}")
    print("=" * 60)


if __name__ == "__main__":
    main()
