#!/usr/bin/env python3
"""
Extract PDF links from crawled pages.

This script:
1. Reads URLs from the Excel output
2. Fetches each page and extracts PDF links
3. Outputs a list of all PDF URLs
"""

import re
import requests
import time
from pathlib import Path
from openpyxl import load_workbook
from urllib.parse import urljoin
from concurrent.futures import ThreadPoolExecutor, as_completed

BASE_DIR = Path(__file__).parent.parent
EXCEL_FILE = BASE_DIR / "mmdc-urls-spider-output.xlsx"
OUTPUT_FILE = BASE_DIR / "pdf-urls.txt"
ALL_URLS_FILE = BASE_DIR / "all-crawled-urls.txt"

# Patterns for document files
DOC_PATTERNS = [
    r'href=["\']([^"\']+\.pdf)["\']',
    r'href=["\']([^"\']+\.PDF)["\']',
    r'src=["\']([^"\']+\.pdf)["\']',
]


def get_urls_from_excel() -> list[str]:
    """Extract all URLs from Excel file."""
    wb = load_workbook(EXCEL_FILE, read_only=True)
    urls = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                urls.append(row[0])

    wb.close()
    return urls


def extract_pdfs_from_url(url: str) -> list[str]:
    """Fetch URL and extract PDF links."""
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()

        pdfs = []
        for pattern in DOC_PATTERNS:
            matches = re.findall(pattern, response.text, re.IGNORECASE)
            for match in matches:
                # Convert relative to absolute URL
                absolute_url = urljoin(url, match)
                if 'mmdc.nl' in absolute_url:
                    pdfs.append(absolute_url)

        return pdfs
    except Exception as e:
        print(f"  Error fetching {url}: {e}")
        return []


def main():
    """Extract all PDF links."""
    print("=" * 60)
    print("PDF Link Extractor")
    print("=" * 60)

    # Get URLs from Excel
    urls = get_urls_from_excel()
    print(f"Loaded {len(urls):,} URLs from Excel")

    # Save all crawled URLs
    ALL_URLS_FILE.write_text("\n".join(urls) + "\n")
    print(f"Saved all URLs to {ALL_URLS_FILE}")

    # Filter to HTML pages (skip static assets)
    html_urls = [u for u in urls if not any(ext in u.lower() for ext in
                 ['.css', '.js', '.png', '.jpg', '.gif', '.svg', '.ico', '.woff', '.ttf'])]
    print(f"Checking {len(html_urls):,} HTML pages for PDF links...")

    all_pdfs = set()

    # Extract PDFs with progress
    for i, url in enumerate(html_urls, 1):
        if i % 20 == 0:
            print(f"Progress: {i}/{len(html_urls)} pages checked, {len(all_pdfs)} PDFs found")

        pdfs = extract_pdfs_from_url(url)
        all_pdfs.update(pdfs)
        time.sleep(0.2)  # Be polite

    # Sort and save
    sorted_pdfs = sorted(all_pdfs)
    OUTPUT_FILE.write_text("\n".join(sorted_pdfs) + "\n")

    print("=" * 60)
    print(f"Found {len(sorted_pdfs):,} unique PDF URLs")
    print(f"Output: {OUTPUT_FILE}")
    print("=" * 60)

    if sorted_pdfs:
        print("\nSample PDFs:")
        for pdf in sorted_pdfs[:10]:
            print(f"  {pdf}")


if __name__ == "__main__":
    main()
