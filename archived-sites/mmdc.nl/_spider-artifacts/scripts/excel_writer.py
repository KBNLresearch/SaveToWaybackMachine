"""Excel writer for mmdc.nl URL Spider with streaming support."""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import SHEET_NAMES, OUTPUT_EXCEL, CHECKPOINTS_DIR


class ExcelURLWriter:
    """Streaming Excel writer for URL data with checkpointing."""

    HEADERS = ["URL", "Path", "Discovered", "Batch", "Status", "Notes"]

    def __init__(self, output_path: Path = OUTPUT_EXCEL):
        self.output_path = output_path
        self.url_buffers: Dict[str, List[dict]] = defaultdict(list)
        self.batch_number = 0
        self.total_urls = 0
        self.sheet_counts: Dict[str, int] = defaultdict(int)
        self._initialized = False

    def initialize(self) -> None:
        """Create or load Excel workbook with all sheets."""
        if self.output_path.exists():
            # Load existing workbook
            self.wb = load_workbook(self.output_path)
            self._load_existing_counts()
        else:
            # Create new workbook
            self.wb = Workbook()

            # Remove default sheet
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]

            # Create all sheets with headers
            for sheet_name in SHEET_NAMES:
                ws = self.wb.create_sheet(sheet_name)
                self._write_headers(ws)

            self.wb.save(self.output_path)

        self._initialized = True

    def _write_headers(self, ws) -> None:
        """Write and style header row."""
        ws.append(self.HEADERS)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Set column widths
        ws.column_dimensions['A'].width = 80  # URL
        ws.column_dimensions['B'].width = 50  # Path
        ws.column_dimensions['C'].width = 20  # Discovered
        ws.column_dimensions['D'].width = 8   # Batch
        ws.column_dimensions['E'].width = 10  # Status
        ws.column_dimensions['F'].width = 30  # Notes

    def _load_existing_counts(self) -> None:
        """Load counts from existing workbook."""
        for sheet_name in self.wb.sheetnames:
            if sheet_name in SHEET_NAMES:
                ws = self.wb[sheet_name]
                # Count rows minus header
                self.sheet_counts[sheet_name] = ws.max_row - 1 if ws.max_row > 1 else 0
                self.total_urls += self.sheet_counts[sheet_name]

    def add_url(self, url: str, sheet_name: str, path: str = "", notes: str = "") -> None:
        """Add URL to buffer for later writing.

        Args:
            url: Full URL
            sheet_name: Target sheet (functional group)
            path: URL path component
            notes: Optional notes
        """
        if not self._initialized:
            self.initialize()

        self.url_buffers[sheet_name].append({
            "url": url,
            "path": path,
            "discovered": datetime.now().isoformat(),
            "batch": self.batch_number,
            "status": "pending",
            "notes": notes,
        })

    def flush(self, force: bool = False) -> int:
        """Write buffered URLs to Excel.

        Args:
            force: Write even if buffer is small

        Returns:
            Number of URLs written
        """
        if not self._initialized:
            self.initialize()

        total_buffered = sum(len(urls) for urls in self.url_buffers.values())

        if total_buffered == 0:
            return 0

        if not force and total_buffered < 100:
            return 0

        written = 0

        for sheet_name, urls in self.url_buffers.items():
            if not urls:
                continue

            if sheet_name not in self.wb.sheetnames:
                ws = self.wb.create_sheet(sheet_name)
                self._write_headers(ws)
            else:
                ws = self.wb[sheet_name]

            for url_data in urls:
                ws.append([
                    url_data["url"],
                    url_data["path"],
                    url_data["discovered"],
                    url_data["batch"],
                    url_data["status"],
                    url_data["notes"],
                ])
                written += 1
                self.sheet_counts[sheet_name] += 1

        self.wb.save(self.output_path)
        self.total_urls += written
        self.batch_number += 1

        # Clear buffers
        self.url_buffers = defaultdict(list)

        # Save checkpoint
        self._save_checkpoint()

        return written

    def _save_checkpoint(self) -> None:
        """Save checkpoint with current state."""
        CHECKPOINTS_DIR.mkdir(parents=True, exist_ok=True)

        checkpoint = {
            "batch_number": self.batch_number,
            "total_urls": self.total_urls,
            "timestamp": datetime.now().isoformat(),
            "sheet_counts": dict(self.sheet_counts),
        }

        checkpoint_file = CHECKPOINTS_DIR / f"checkpoint_{self.batch_number:04d}.json"
        with open(checkpoint_file, "w") as f:
            json.dump(checkpoint, f, indent=2)

    def get_stats(self) -> Dict:
        """Get current statistics."""
        return {
            "total_urls": self.total_urls,
            "batch_number": self.batch_number,
            "sheet_counts": dict(self.sheet_counts),
            "buffered": sum(len(urls) for urls in self.url_buffers.values()),
        }

    def close(self) -> None:
        """Flush remaining data and close workbook."""
        self.flush(force=True)
        if hasattr(self, 'wb'):
            self.wb.close()
