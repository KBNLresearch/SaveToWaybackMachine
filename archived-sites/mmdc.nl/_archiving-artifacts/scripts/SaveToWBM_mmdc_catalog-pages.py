#!/usr/bin/env python
"""
Save mmdc.nl CATALOG PAGE URLs to Wayback Machine.

This script:
1. Reads catalog page URLs from mmdc-urls-UNIFIED.xlsx, sheet 'catalog-pages'
2. Submits each URL to Wayback Machine using authenticated SPN2 API
3. Updates the Excel with WBM results (columns D, E, F)
4. Maintains full progress state for resume after crashes/interruptions

Total: 11,738 catalog page URLs. Estimated runtime: ~55 hours at 17s/request.

Excel column layout (catalog-pages sheet):
- Col A (1):  URL-new-temp (formula, constructed from col L)
- Col B (2):  Hyperlink
- Col C (3):  URLnewTemp_status200-OK
- Col D (4):  WBM_URL          <-- written by this script
- Col E (5):  WBM_Timestamp    <-- written by this script
- Col F (6):  WBM_HTTP_Status  <-- written by this script
- Col G (7):  URL-old-defunct
- Col H (8):  Section
- Col I (9):  Title
- Col J (10): Notes
- Col K (11): Archive_Status
- Col L (12): Local_File (e.g. catalog-page-2.html)
- Col M (13): Local_Path_Absolute
- Col N (14): Local_Path_Relative
- Col O (15): Local_File_Exists

URL construction:
  https://mmdc.nl/wbm/site/search/{Local_File}
  (Col A contains this as a formula, but we construct it from Col L to avoid
   formula evaluation issues with openpyxl)

ROBUSTNESS FEATURES (designed for 55+ hour runs with interruptions):
- Resume capability: progress saved after EVERY URL, picks up exactly where it left off
- Graceful shutdown: Ctrl+C saves Excel + progress before exit
- Pause file: create PAUSE.flag next to this script to pause after current URL
- Offline detection: after 3 consecutive transient failures, enters "wait for
  connectivity" mode with exponential backoff (30s -> 1m -> 2m -> 5m -> 10m)
  instead of burning through URLs marking them all as failed
- Connectivity ping: tests web.archive.org reachability before resuming after offline
- Transient vs permanent failure tracking:
    - Transient (timeout, connection error, rate limit): row left empty, retried
      automatically on next run
    - Permanent (HTTP 523 origin unreachable, etc.): marked FAILED in Excel col F
- Excel saved after EVERY successful submission to minimize data loss on crash
- Excel file locking detection with retry (in case file is open in Excel)
- Rate limit handling with 5-minute pause + jitter

LOGGING (3 output files):
- _archiving-artifacts/logs/catalog-archive-STATUS.txt
    Human-readable status snapshot, overwritten on every URL. Open this file
    at any time to check current progress! Shows progress bar, rates, ETA,
    session history. Example:
        Progress:  [########--------------------------------] 20.1%
        Archived:  2360 / 11738
        Rate:      182.3 URLs/hour
        ETA:       51.4 hours (est. completion: 2026-04-04 18:00)

- _archiving-artifacts/logs/catalog-archive-log.txt
    Detailed append-only per-URL log with timestamps, row numbers, job IDs,
    running totals, offline/online events, and session boundaries.

- _archiving-artifacts/data/catalog-archive-progress.json
    Machine-readable progress state, also used for resume capability. Contains:
        - submitted / failed_permanent / failed_transient / skipped counts
        - rate_urls_per_hour (computed live)
        - eta_remaining (e.g. "32.5 hours (est. completion: 2026-04-04 15:30)")
        - last_url, last_result, last_activity timestamp
        - Full session history (start/end/duration/counts per run)

USAGE:
    python SaveToWBM_mmdc_catalog.py                  # Start or resume
    python SaveToWBM_mmdc_catalog.py --retry           # Reset row position to re-scan all rows
                                                       # (already-archived rows are still skipped
                                                       #  based on valid WBM_URL in col D)
    python SaveToWBM_mmdc_catalog.py --limit 5         # Process only N URLs (for testing)
    python SaveToWBM_mmdc_catalog.py --from 100 --to 200
                                                       # Process only data rows 100-200
                                                       # (row 1 = first data row in Excel)
                                                       # --from and --to can be used independently

CONTROLS WHILE RUNNING:
    Ctrl+C          -> Graceful shutdown (saves Excel + progress + status, then exits)
    Create PAUSE.flag file next to this script
                    -> Pause after current URL (saves state, exits)
                       Delete the flag file before running again.

CREDENTIALS:
    Requires .env file in the same folder as this script with:
        IA_ACCESS_KEY=your_access_key
        IA_SECRET_KEY=your_secret_key

To fully reset and start fresh, delete BOTH:
    _archiving-artifacts/data/catalog-archive-progress.json
    AND clear columns D, E, F in the Excel
"""

import sys
import time
import json
import signal
import requests
import random
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# Configuration
SCRIPT_DIR = Path(__file__).parent
ARTIFACTS_DIR = SCRIPT_DIR.parent
DATA_DIR = ARTIFACTS_DIR / "data"
LOGS_DIR = ARTIFACTS_DIR / "logs"
MMDC_DIR = ARTIFACTS_DIR.parent  # mmdc.nl folder

# Input/Output files
EXCEL_FILE = MMDC_DIR / "mmdc-urls-UNIFIED.xlsx"
SHEET_NAME = "catalog-pages"
PROGRESS_FILE = DATA_DIR / "catalog-archive-progress.json"
LOG_FILE = LOGS_DIR / "catalog-archive-log.txt"
STATUS_FILE = LOGS_DIR / "catalog-archive-STATUS.txt"
PAUSE_FILE = SCRIPT_DIR / "PAUSE.flag"

# Credentials
ENV_FILE = SCRIPT_DIR / ".env"

# URL construction
BASE_URL = "https://mmdc.nl/wbm/site/search/"

# Column indices (1-based) - updated to match current Excel layout
COL_URL_TEMP = 1       # A: URL-new-temp (formula, read-only)
COL_STATUS_200 = 3     # C: URLnewTemp_status200-OK
COL_WBM_URL = 4        # D: WBM_URL
COL_WBM_TIMESTAMP = 5  # E: WBM_Timestamp
COL_WBM_HTTP = 8       # H: WBM_HTTP_Status
COL_LOCAL_FILE = 14    # N: Local_File (used to construct URL)

# Rate limiting
SPN_DELAY = 17         # seconds between SPN requests (15 req/min limit)
MAX_RETRIES = 3        # retries per individual URL (before declaring it failed for this attempt)
EXCEL_SAVE_RETRY = 5
RATE_LIMIT_WAIT = 300  # 5 minutes on rate limit

# Offline detection
CONSECUTIVE_FAIL_THRESHOLD = 3  # after this many consecutive failures, assume offline
OFFLINE_BACKOFF_STEPS = [30, 60, 120, 300, 600]  # seconds: 30s, 1m, 2m, 5m, 10m (then repeat 10m)
CONNECTIVITY_CHECK_URL = "https://web.archive.org"
CONNECTIVITY_CHECK_TIMEOUT = 15

# User-Agent
USER_AGENT = "KB-Archiver/1.0 (mmdc.nl catalog archiving project; Contact: kb.nl)"

# Color fills for Excel
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Global state for graceful shutdown
SHUTDOWN_REQUESTED = False
CURRENT_WORKBOOK = None
CURRENT_PROGRESS = None


# ---------------------------------------------------------------------------
#  Signal handling
# ---------------------------------------------------------------------------

def signal_handler(signum, frame):
    """Handle Ctrl+C gracefully - save state before exit."""
    global SHUTDOWN_REQUESTED
    print("\n\n" + "!" * 70)
    print("SHUTDOWN REQUESTED - Saving state...")
    print("!" * 70)
    SHUTDOWN_REQUESTED = True

    try:
        if CURRENT_PROGRESS:
            save_progress(CURRENT_PROGRESS)
            print("  Progress saved to JSON")
        if CURRENT_WORKBOOK:
            save_excel_safely(CURRENT_WORKBOOK, EXCEL_FILE)
            print("  Excel file saved")
        if CURRENT_PROGRESS:
            update_status_file(CURRENT_PROGRESS, "INTERRUPTED (Ctrl+C)")
            print("  Status file updated")
    except Exception as e:
        print(f"  Warning: Could not save all state: {e}")

    print("\nYou can resume by running the script again.")
    print(f"To start fresh, delete: {PROGRESS_FILE}")
    sys.exit(0)


signal.signal(signal.SIGINT, signal_handler)
if hasattr(signal, 'SIGTERM'):
    signal.signal(signal.SIGTERM, signal_handler)


# ---------------------------------------------------------------------------
#  Credentials
# ---------------------------------------------------------------------------

def load_credentials():
    """Load IA S3 credentials from .env file."""
    if not ENV_FILE.exists():
        raise FileNotFoundError(
            f".env file not found at {ENV_FILE}\n"
            f"Create it with:\n"
            f"  IA_ACCESS_KEY=your_access_key\n"
            f"  IA_SECRET_KEY=your_secret_key"
        )

    creds = {}
    with open(ENV_FILE, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                creds[key.strip()] = value.strip().strip("'\"")

    access_key = creds.get('IA_ACCESS_KEY')
    secret_key = creds.get('IA_SECRET_KEY')

    if not access_key or not secret_key:
        raise ValueError("IA_ACCESS_KEY and IA_SECRET_KEY must be set in .env file")

    return access_key, secret_key


# ---------------------------------------------------------------------------
#  Progress & state persistence
# ---------------------------------------------------------------------------

def load_progress():
    """Load progress for resume capability."""
    if PROGRESS_FILE.exists():
        try:
            with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"  Warning: Could not load progress file: {e}")

    return {
        'start_time': None,
        'last_row': 1,          # Last processed Excel row (1 = header, so none processed)
        'total_in_sheet': 0,    # Total data rows in the sheet
        'total_processed': 0,
        'submitted': 0,
        'failed_permanent': 0,  # Truly failed (e.g. HTTP 523 origin unreachable)
        'failed_transient': 0,  # Failed due to connectivity/timeout (will retry on next run)
        'skipped': 0,           # Already had WBM_URL
        'last_url': None,
        'last_result': None,
        'last_activity': None,
        'rate_urls_per_hour': 0,
        'eta_remaining': None,
        'sessions': [],         # Log of start/stop times
    }


def save_progress(progress):
    """Save progress JSON with computed rates and ETA."""
    try:
        # Compute rate and ETA
        if progress['start_time'] and progress['submitted'] > 0:
            start = datetime.fromisoformat(progress['start_time'])
            elapsed = (datetime.now() - start).total_seconds()
            if elapsed > 0:
                progress['rate_urls_per_hour'] = round(progress['submitted'] / elapsed * 3600, 1)
                remaining = progress['total_in_sheet'] - progress['submitted'] - progress['skipped'] - progress['failed_permanent']
                if progress['rate_urls_per_hour'] > 0:
                    eta_seconds = remaining / progress['rate_urls_per_hour'] * 3600
                    eta_dt = datetime.now() + timedelta(seconds=eta_seconds)
                    progress['eta_remaining'] = f"{eta_seconds/3600:.1f} hours (est. completion: {eta_dt.strftime('%Y-%m-%d %H:%M')})"

        progress['last_activity'] = datetime.now().isoformat()

        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
            json.dump(progress, f, indent=2)
    except Exception as e:
        print(f"  Warning: Could not save progress: {e}")


def update_status_file(progress, state="RUNNING"):
    """Write a human-readable status snapshot.

    This file is designed to be opened at any time to see where we are.
    It gets overwritten on every update (not appended).
    """
    try:
        LOGS_DIR.mkdir(parents=True, exist_ok=True)

        archived = progress['submitted'] + progress['skipped']
        total = progress['total_in_sheet']
        remaining = total - archived - progress['failed_permanent']
        pct = (archived * 100) / total if total > 0 else 0
        bar_len = 40
        bar_filled = int(bar_len * pct / 100)
        bar = "#" * bar_filled + "-" * (bar_len - bar_filled)

        lines = [
            "=" * 60,
            "  MMDC.nl Catalog WBM Archiver - STATUS",
            "=" * 60,
            "",
            f"  State:           {state}",
            f"  Last updated:    {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"  Running since:   {progress['start_time'] or 'N/A'}",
            "",
            f"  Progress:        [{bar}] {pct:.1f}%",
            f"  Archived:        {archived} / {total}",
            "",
            f"    Submitted (this + prev runs):  {progress['submitted']}",
            f"    Skipped (already had WBM_URL): {progress['skipped']}",
            f"    Failed (permanent):            {progress['failed_permanent']}",
            f"    Failed (transient, will retry): {progress['failed_transient']}",
            f"    Remaining:                     {remaining}",
            "",
            f"  Current position: row {progress['last_row']} of {total + 1}",
            f"  Last URL:         {progress.get('last_url', 'N/A')}",
            f"  Last result:      {progress.get('last_result', 'N/A')}",
            "",
            f"  Rate:            {progress.get('rate_urls_per_hour', 0):.1f} URLs/hour",
            f"  ETA:             {progress.get('eta_remaining', 'calculating...')}",
            "",
            "-" * 60,
            f"  Sessions: {len(progress.get('sessions', []))}",
        ]

        for i, session in enumerate(progress.get('sessions', [])[-5:], 1):
            started = session.get('started', '?')[:19]
            ended = session.get('ended', 'ongoing')[:19]
            mode = session.get('mode', '?')
            sub = session.get('submitted', '?')
            fail = session.get('failed', '?')
            lines.append(f"    {i}. {started} -> {ended} [{mode}] sub={sub} fail={fail}")

        lines += [
            "",
            "-" * 60,
            "  Files:",
            f"    Excel:    {EXCEL_FILE}",
            f"    Progress: {PROGRESS_FILE}",
            f"    Log:      {LOG_FILE}",
            f"    Status:   {STATUS_FILE}  (this file)",
            "",
            "  Controls:",
            "    Ctrl+C         -> Graceful shutdown",
            f"    Create {PAUSE_FILE.name}  -> Pause after current URL",
            "=" * 60,
        ]

        with open(STATUS_FILE, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines) + '\n')
    except Exception:
        pass  # Status file is best-effort, never crash for it


def save_excel_safely(wb, filepath, max_retries=EXCEL_SAVE_RETRY):
    """Save Excel with retry logic for file locking."""
    for attempt in range(max_retries):
        try:
            wb.save(filepath)
            return True
        except PermissionError:
            if attempt < max_retries - 1:
                wait = 5 * (attempt + 1)
                print(f"\n  Excel file locked, waiting {wait}s... (close Excel if open)")
                time.sleep(wait)
            else:
                print(f"\n  ERROR: Could not save Excel after {max_retries} attempts")
                print(f"  Please close the file and run the script again")
                return False
        except Exception as e:
            print(f"\n  Error saving Excel: {e}")
            return False
    return False


# ---------------------------------------------------------------------------
#  Connectivity
# ---------------------------------------------------------------------------

def check_connectivity():
    """Check if web.archive.org is reachable."""
    try:
        r = requests.head(CONNECTIVITY_CHECK_URL, timeout=CONNECTIVITY_CHECK_TIMEOUT)
        return r.status_code < 500
    except Exception:
        return False


def wait_for_connectivity(log):
    """Wait until internet connectivity is restored. Returns True when online, False on shutdown/pause."""
    global SHUTDOWN_REQUESTED

    msg = (f"\n{'=' * 70}\n"
           f"OFFLINE DETECTED - Waiting for connectivity to return...\n"
           f"  (Ctrl+C to save and exit, or create PAUSE.flag to pause)\n"
           f"{'=' * 70}")
    print(msg)
    log.write(f"\n--- OFFLINE at {datetime.now().isoformat()} ---\n")
    log.flush()

    if CURRENT_PROGRESS:
        update_status_file(CURRENT_PROGRESS, "OFFLINE - waiting for connectivity")

    backoff_idx = 0
    total_waited = 0

    while not SHUTDOWN_REQUESTED:
        # Check for pause file
        if PAUSE_FILE.exists():
            print("\n  PAUSE.flag detected while waiting for connectivity.")
            return False

        wait_time = OFFLINE_BACKOFF_STEPS[min(backoff_idx, len(OFFLINE_BACKOFF_STEPS) - 1)]
        waited_min = total_waited / 60
        print(f"\n  Waiting {wait_time}s before checking... (total offline: {waited_min:.1f} min)", end="", flush=True)
        time.sleep(wait_time)
        total_waited += wait_time

        print(" pinging web.archive.org...", end="", flush=True)
        if check_connectivity():
            print(" ONLINE!")
            print("=" * 70)
            print("Connectivity restored! Resuming archiving...\n")
            log.write(f"--- ONLINE at {datetime.now().isoformat()} (was offline {waited_min:.1f} min) ---\n\n")
            log.flush()
            if CURRENT_PROGRESS:
                update_status_file(CURRENT_PROGRESS, "RUNNING (recovered from offline)")
            return True

        print(" still offline.", end="", flush=True)
        backoff_idx += 1
        if CURRENT_PROGRESS:
            update_status_file(CURRENT_PROGRESS, f"OFFLINE - waiting ({waited_min:.1f} min so far)")

    return False


# ---------------------------------------------------------------------------
#  WBM submission
# ---------------------------------------------------------------------------

def submit_to_wbm(url, access_key, secret_key):
    """Submit URL to Wayback Machine SPN2 API with retry logic.

    Returns dict with:
        success: bool
        transient: bool (True if failure is likely due to connectivity, not permanent)
        + job_id, wbm_url, timestamp, message, http_status on success
        + error on failure
    """
    headers = {
        'Accept': 'application/json',
        'Authorization': f'LOW {access_key}:{secret_key}',
        'User-Agent': USER_AGENT
    }

    for attempt in range(MAX_RETRIES):
        try:
            response = requests.post(
                'https://web.archive.org/save',
                headers=headers,
                data={'url': url},
                timeout=180
            )

            if response.status_code == 200:
                result = response.json()
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                return {
                    'success': True,
                    'transient': False,
                    'job_id': result.get('job_id', ''),
                    'wbm_url': f"https://web.archive.org/web/{timestamp}/{url}",
                    'timestamp': timestamp,
                    'message': result.get('message', ''),
                    'http_status': 200
                }
            elif response.status_code == 429:
                if attempt < MAX_RETRIES - 1:
                    wait = RATE_LIMIT_WAIT + random.uniform(30, 120)
                    print(f"\n    RATE LIMITED! Waiting {wait:.0f}s...", end="", flush=True)
                    time.sleep(wait)
                    continue
                return {'success': False, 'transient': True, 'error': 'Rate limited after retries'}
            elif response.status_code == 523:
                # Origin unreachable - this is a permanent issue (site/page problem)
                return {'success': False, 'transient': False, 'error': 'HTTP 523 Origin Unreachable'}
            elif response.status_code >= 500:
                if attempt < MAX_RETRIES - 1:
                    wait = 30 * (attempt + 1)
                    print(f"\n    Server error {response.status_code}, retry in {wait}s...", end="", flush=True)
                    time.sleep(wait)
                    continue
                return {'success': False, 'transient': True, 'error': f'HTTP {response.status_code} after retries'}
            else:
                return {'success': False, 'transient': False, 'error': f'HTTP {response.status_code}'}

        except requests.exceptions.Timeout:
            if attempt < MAX_RETRIES - 1:
                wait = 30 * (attempt + 1)
                print(f"\n    Timeout, retry in {wait}s...", end="", flush=True)
                time.sleep(wait)
            else:
                return {'success': False, 'transient': True, 'error': 'Timeout after retries'}
        except requests.exceptions.ConnectionError as e:
            if attempt < MAX_RETRIES - 1:
                wait = 30 * (attempt + 1)
                print(f"\n    Connection error, retry in {wait}s...", end="", flush=True)
                time.sleep(wait)
            else:
                return {'success': False, 'transient': True, 'error': f'Connection error: {str(e)[:80]}'}
        except Exception as e:
            if attempt < MAX_RETRIES - 1:
                time.sleep(15 * (attempt + 1))
            else:
                return {'success': False, 'transient': True, 'error': str(e)[:100]}

    return {'success': False, 'transient': True, 'error': 'Max retries exceeded'}


# ---------------------------------------------------------------------------
#  Row eligibility
# ---------------------------------------------------------------------------

def should_retry_row(ws, row):
    """Check if a row needs (re)submission.
    Returns True if the row has no successful WBM archive yet.
    Rows with a valid WBM_URL are considered done.
    Rows with empty, None, or FAILED status are eligible for retry.
    """
    wbm_url = ws.cell(row=row, column=COL_WBM_URL).value
    if wbm_url and isinstance(wbm_url, str) and wbm_url.startswith("https://web.archive.org/"):
        return False
    return True


# ---------------------------------------------------------------------------
#  Main
# ---------------------------------------------------------------------------

def main():
    global CURRENT_WORKBOOK, CURRENT_PROGRESS, SHUTDOWN_REQUESTED

    retry_mode = '--retry' in sys.argv
    limit = None
    if '--limit' in sys.argv:
        try:
            limit = int(sys.argv[sys.argv.index('--limit') + 1])
        except (IndexError, ValueError):
            print("ERROR: --limit requires a number (e.g. --limit 5)")
            sys.exit(1)

    # --from and --to: restrict processing to a range of Excel rows (1-based, data rows)
    # e.g. --from 100 --to 200 processes data rows 100-200 (Excel rows 101-201)
    row_from = None
    row_to = None
    if '--from' in sys.argv:
        try:
            row_from = int(sys.argv[sys.argv.index('--from') + 1])
        except (IndexError, ValueError):
            print("ERROR: --from requires a number (e.g. --from 100)")
            sys.exit(1)
    if '--to' in sys.argv:
        try:
            row_to = int(sys.argv[sys.argv.index('--to') + 1])
        except (IndexError, ValueError):
            print("ERROR: --to requires a number (e.g. --to 200)")
            sys.exit(1)

    print("=" * 70)
    print("MMDC.nl Catalog Pages - Wayback Machine Archiver")
    print("SPN2 API (Authenticated)")
    print("=" * 70)
    print()
    print("CONTROLS:")
    print("  Ctrl+C             -> Graceful shutdown (saves state)")
    print(f"  Create PAUSE.flag  -> Pause after current URL")
    print(f"                        ({PAUSE_FILE})")
    print()
    print("LOGGING:")
    print(f"  Status (open anytime): {STATUS_FILE}")
    print(f"  Detailed log:          {LOG_FILE}")
    print(f"  Progress JSON:         {PROGRESS_FILE}")
    print()
    print("RESILIENCE: Auto-retry | Offline detection & wait |")
    print("            Resume from exact position | Excel saved every URL")
    if limit:
        print(f"\n  *** TEST MODE: --limit {limit} (will stop after {limit} submissions) ***")
    if row_from or row_to:
        f = row_from or 1
        t = row_to or '(end)'
        print(f"\n  *** RANGE MODE: processing data rows {f} to {t} ***")
    print("=" * 70)
    print()

    # Clean up pause file from previous run
    if PAUSE_FILE.exists():
        PAUSE_FILE.unlink()
        print("  (Removed leftover PAUSE.flag from previous run)")

    # Ensure directories exist
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)

    # Check Excel file
    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    # Load credentials
    print("Loading credentials...")
    try:
        access_key, secret_key = load_credentials()
        print(f"  Access key: {access_key[:4]}...{access_key[-4:]}")
        print("  Credentials loaded successfully")
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    # Load Excel
    print(f"\nLoading {EXCEL_FILE}...")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        CURRENT_WORKBOOK = wb
    except Exception as e:
        print(f"ERROR: Could not load Excel: {e}")
        sys.exit(1)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    total_rows = ws.max_row - 1  # Exclude header
    print(f"  Sheet '{SHEET_NAME}': {total_rows} URLs")

    # Verify column layout
    expected_headers = {
        COL_URL_TEMP: 'URL-new-temp',
        COL_STATUS_200: 'URLnewTemp_status200-OK',
        COL_WBM_URL: 'WBM_URL',
        COL_WBM_TIMESTAMP: 'WBM_Timestamp',
        COL_WBM_HTTP: 'WBM_HTTP_Status',
        COL_LOCAL_FILE: 'Local_File',
    }
    print("\n  Verifying column layout...")
    layout_ok = True
    for col, expected in expected_headers.items():
        actual = ws.cell(row=1, column=col).value
        match = "OK" if actual == expected else "MISMATCH"
        if actual != expected:
            layout_ok = False
            print(f"    Col {col}: expected '{expected}', got '{actual}' *** {match} ***")
        else:
            print(f"    Col {col}: {actual} - {match}")

    if not layout_ok:
        print("\n  ERROR: Column layout does not match expected format!")
        print("  Please check the Excel file and update column indices in this script.")
        sys.exit(1)

    print("  Column layout verified!")

    # Pre-scan: count how many rows already have a valid WBM_URL vs need work
    already_done = 0
    needs_work = 0
    for row in range(2, ws.max_row + 1):
        if should_retry_row(ws, row):
            needs_work += 1
        else:
            already_done += 1

    print(f"\n  Pre-scan results:")
    print(f"    Already archived (valid WBM_URL): {already_done}")
    print(f"    Needs submission (new or retry):   {needs_work}")

    if needs_work == 0:
        print("\n  All URLs already have a valid WBM_URL. Nothing to do!")
        print("  To force re-archive, clear columns D-F in the Excel and run again.")
        sys.exit(0)

    # Load progress
    progress = load_progress()
    progress['total_in_sheet'] = total_rows
    CURRENT_PROGRESS = progress

    if retry_mode:
        print("\n  --retry flag: Resetting row position to scan all rows.")
        print("  (Already-archived URLs will still be skipped based on WBM_URL in col D)")
        progress['last_row'] = 1

    if progress['start_time'] and progress['last_row'] > 1 and not retry_mode:
        rows_remaining_from_position = sum(
            1 for r in range(progress['last_row'] + 1, ws.max_row + 1)
            if should_retry_row(ws, r)
        )
        print(f"\n  RESUMING from previous run:")
        print(f"    First started: {progress['start_time']}")
        print(f"    Last row processed: {progress['last_row']}")
        print(f"    Submitted so far: {progress['submitted']}")
        print(f"    Failed (permanent): {progress['failed_permanent']}")
        print(f"    Failed (transient, will retry): {progress['failed_transient']}")
        print(f"    Skipped (already archived): {progress['skipped']}")
        print(f"    Remaining from current position: {rows_remaining_from_position}")
        needs_work = rows_remaining_from_position
    else:
        if not progress['start_time']:
            progress['start_time'] = datetime.now().isoformat()

    # Log this session
    progress['sessions'].append({
        'started': datetime.now().isoformat(),
        'resumed_from_row': progress['last_row'],
        'mode': 'retry' if retry_mode else 'resume' if progress['last_row'] > 1 else 'fresh'
    })

    # Time estimate
    est_hours = (needs_work * (SPN_DELAY + 5)) / 3600
    print(f"\n  Estimated time for {needs_work} URLs: ~{est_hours:.1f} hours")
    print(f"  Delay between requests: {SPN_DELAY}s + random jitter")

    # Write initial status file
    update_status_file(progress, "STARTING")

    print("\n  Starting in 3 seconds... (Ctrl+C to abort)")
    time.sleep(3)

    print("\n" + "-" * 70)
    print("Processing catalog pages...")
    print("-" * 70 + "\n")

    # Open log file
    log_is_new = not LOG_FILE.exists() or (progress['last_row'] <= 1 and not retry_mode)
    with open(LOG_FILE, 'a', encoding='utf-8') as log:
        if log_is_new:
            log.write(f"{'=' * 70}\n")
            log.write(f"MMDC.nl Catalog Pages - WBM Archiving Log\n")
            log.write(f"Started: {datetime.now().isoformat()}\n")
            log.write(f"Total URLs: {total_rows}\n")
            log.write(f"Already archived: {already_done}\n")
            log.write(f"To submit: {needs_work}\n")
            log.write(f"{'=' * 70}\n\n")
        else:
            log.write(f"\n{'=' * 70}\n")
            log.write(f"Session resumed at {datetime.now().isoformat()}\n")
            log.write(f"Resuming from row {progress['last_row'] + 1}\n")
            log.write(f"Submitted so far: {progress['submitted']}\n")
            log.write(f"{'=' * 70}\n\n")

        consecutive_failures = 0
        submitted_this_session = 0
        failed_this_session = 0
        session_start = datetime.now()

        # Determine row range (Excel rows: 2 = first data row = data row 1)
        loop_start = (row_from + 1) if row_from else 2           # Excel row
        loop_end = (row_to + 1) if row_to else ws.max_row        # Excel row (inclusive)

        for row in range(loop_start, loop_end + 1):
            if SHUTDOWN_REQUESTED:
                break

            # --limit: stop after N successful submissions this session
            if limit and submitted_this_session >= limit:
                print(f"\n  --limit {limit} reached. Stopping.")
                log.write(f"\n--- LIMIT {limit} reached at {datetime.now().isoformat()} ---\n")
                log.flush()
                break

            # Check for pause file
            if PAUSE_FILE.exists():
                print(f"\n  PAUSE.flag detected! Saving state and pausing...")
                log.write(f"\n--- PAUSED at {datetime.now().isoformat()} (row {row}) ---\n")
                log.flush()
                update_status_file(progress, "PAUSED (PAUSE.flag)")
                break

            # Skip rows before our resume position
            if row <= progress['last_row'] and progress['last_row'] > 1:
                continue

            # Get local filename from col L, construct URL
            local_file = ws.cell(row=row, column=COL_LOCAL_FILE).value
            if not local_file:
                print(f"  Row {row}: No Local_File value, skipping")
                progress['last_row'] = row
                save_progress(progress)
                continue

            url = BASE_URL + local_file

            # Check if already archived (valid WBM_URL in col D)
            if not should_retry_row(ws, row):
                progress['skipped'] += 1
                progress['last_row'] = row
                progress['total_processed'] += 1
                if progress['total_processed'] % 500 == 0:
                    save_progress(progress)
                    update_status_file(progress, "RUNNING")
                continue

            # Progress display
            progress['total_processed'] += 1
            total_done = progress['submitted'] + progress['skipped']
            pct = (total_done * 100) / total_rows if total_rows > 0 else 0

            # Calculate session rate
            session_elapsed = (datetime.now() - session_start).total_seconds()
            if session_elapsed > 0 and submitted_this_session > 0:
                session_rate = submitted_this_session / session_elapsed * 3600
                rate_str = f"{session_rate:.0f}/hr"
            else:
                rate_str = "..."

            ts = datetime.now().strftime('%H:%M:%S')
            print(f"  {ts} [{total_done}/{total_rows}] ({pct:.1f}%) Row {row}: {local_file}",
                  end=" ", flush=True)

            # Submit to WBM
            result = submit_to_wbm(url, access_key, secret_key)

            if result['success']:
                job_id = result.get('job_id', '')
                job_short = job_id[:16] if len(job_id) > 16 else job_id
                print(f"OK (job: {job_short}) [{rate_str}]")

                # Write results to Excel
                ws.cell(row=row, column=COL_WBM_URL).value = result['wbm_url']
                ws.cell(row=row, column=COL_WBM_TIMESTAMP).value = result['timestamp']
                ws.cell(row=row, column=COL_WBM_HTTP).value = result.get('http_status', 200)

                progress['submitted'] += 1
                progress['last_url'] = url
                progress['last_result'] = f"OK (job: {job_id})"
                submitted_this_session += 1
                consecutive_failures = 0  # Reset failure counter

                # Detailed log entry
                log.write(f"{ts} OK row={row} | {url} | job={job_id} | "
                          f"total={progress['submitted']}/{total_rows}\n")
                log.flush()

                # Save Excel after EVERY successful submission
                if not save_excel_safely(wb, EXCEL_FILE):
                    print("    WARNING: Excel save failed! Data is in memory, will retry next save.")

            else:
                error = result.get('error', 'Unknown')
                is_transient = result.get('transient', True)
                fail_type = 'transient' if is_transient else 'PERMANENT'
                print(f"FAILED ({fail_type}): {error}")

                if is_transient:
                    # Don't write anything to Excel for transient failures.
                    # The row stays "empty" so it will be retried on the next pass.
                    progress['failed_transient'] += 1
                else:
                    # Permanent failure: mark in Excel so we know it was attempted
                    ws.cell(row=row, column=COL_WBM_URL).value = None
                    ws.cell(row=row, column=COL_WBM_TIMESTAMP).value = None
                    ws.cell(row=row, column=COL_WBM_HTTP).value = f'FAILED: {error[:40]}'
                    ws.cell(row=row, column=COL_WBM_HTTP).fill = RED_FILL
                    progress['failed_permanent'] += 1
                    save_excel_safely(wb, EXCEL_FILE)

                progress['last_url'] = url
                progress['last_result'] = f"FAILED ({fail_type}): {error}"
                failed_this_session += 1
                consecutive_failures += 1

                log.write(f"{ts} FAILED ({fail_type}) row={row} | {url} | {error}\n")
                log.flush()

                # OFFLINE DETECTION: if multiple consecutive failures, we're probably offline
                if consecutive_failures >= CONSECUTIVE_FAIL_THRESHOLD and is_transient:
                    log.write(f"\n--- OFFLINE DETECTED at {datetime.now().isoformat()} "
                              f"({consecutive_failures} consecutive transient failures) ---\n")
                    log.flush()

                    # Save state before waiting, back up last_row so failed rows get retried
                    progress['last_row'] = row - consecutive_failures
                    save_progress(progress)
                    save_excel_safely(wb, EXCEL_FILE)
                    update_status_file(progress, "OFFLINE - waiting for connectivity")

                    # Wait for connectivity
                    if wait_for_connectivity(log):
                        consecutive_failures = 0
                        # Don't advance last_row - the failed rows will be retried
                        continue
                    else:
                        # Shutdown or pause requested during wait
                        break

            # Update progress and status file
            progress['last_row'] = row
            save_progress(progress)
            update_status_file(progress, "RUNNING")

            # Rate limiting delay (skip for last URL)
            if row < ws.max_row and not SHUTDOWN_REQUESTED:
                delay = SPN_DELAY + random.uniform(2, 8)
                time.sleep(delay)

            # Periodic summary in log and console
            if (submitted_this_session + failed_this_session) % 50 == 0 and (submitted_this_session + failed_this_session) > 0:
                elapsed_str = str(timedelta(seconds=int(session_elapsed)))
                summary = (f"\n    === MILESTONE ({submitted_this_session + failed_this_session} this session, "
                           f"elapsed {elapsed_str}) ===\n"
                           f"    Session: {submitted_this_session} submitted, {failed_this_session} failed\n"
                           f"    Overall: {progress['submitted']}/{total_rows} archived, "
                           f"{progress['failed_permanent']} perm-failed, "
                           f"{progress['failed_transient']} transient | {rate_str}\n")
                print(summary)
                log.write(summary + "\n")
                log.flush()

        # Update session end time
        if progress['sessions']:
            progress['sessions'][-1]['ended'] = datetime.now().isoformat()
            progress['sessions'][-1]['submitted'] = submitted_this_session
            progress['sessions'][-1]['failed'] = failed_this_session
            session_elapsed = (datetime.now() - session_start).total_seconds()
            progress['sessions'][-1]['duration_minutes'] = round(session_elapsed / 60, 1)

    # Final save
    print("\n" + "-" * 70)
    print("Saving final state...")
    save_excel_safely(wb, EXCEL_FILE)
    progress['last_activity'] = datetime.now().isoformat()
    save_progress(progress)

    # Count actual state of Excel
    final_done = sum(1 for r in range(2, ws.max_row + 1) if not should_retry_row(ws, r))
    final_remaining = total_rows - final_done

    # Determine final state
    paused = PAUSE_FILE.exists()
    if SHUTDOWN_REQUESTED:
        final_state = "INTERRUPTED (Ctrl+C) - run again to resume"
    elif paused:
        final_state = "PAUSED (PAUSE.flag) - delete flag and run again"
    elif final_remaining == 0:
        final_state = "COMPLETE - all URLs archived!"
    else:
        final_state = f"SESSION ENDED - {final_remaining} URLs remaining"

    update_status_file(progress, final_state)

    # Summary
    print()
    print("=" * 70)
    print(f"  {final_state}")
    print("=" * 70)
    print(f"\n  This session:")
    print(f"    Submitted: {submitted_this_session}")
    print(f"    Failed: {failed_this_session}")
    session_elapsed = (datetime.now() - session_start).total_seconds()
    print(f"    Duration: {timedelta(seconds=int(session_elapsed))}")
    print(f"\n  Overall progress:")
    print(f"    Archived (valid WBM_URL): {final_done} / {total_rows}")
    print(f"    Remaining: {final_remaining}")
    print(f"    Total submitted (all sessions): {progress['submitted']}")
    print(f"    Total failed permanent: {progress['failed_permanent']}")
    print(f"    Total failed transient: {progress['failed_transient']}")
    print()
    print(f"  Files:")
    print(f"    Excel:    {EXCEL_FILE}")
    print(f"    Progress: {PROGRESS_FILE}")
    print(f"    Log:      {LOG_FILE}")
    print(f"    Status:   {STATUS_FILE}  <-- open anytime to check progress")

    if final_remaining > 0:
        print(f"\n  Next steps:")
        if progress['failed_transient'] > 0:
            print(f"    - {progress['failed_transient']} transient failures will be retried automatically")
        if progress['failed_permanent'] > 0:
            print(f"    - {progress['failed_permanent']} permanent failures (check log for details)")
        print(f"    - Run this script again to continue from row {progress['last_row'] + 1}")
        print(f"    - Or run with --retry to re-scan all rows (already-archived rows still skipped)")
    else:
        print(f"\n  All {total_rows} catalog pages archived!")

    if paused:
        print(f"\n  Note: PAUSE.flag still exists at {PAUSE_FILE}")
        print(f"  Delete it before running again.")


if __name__ == "__main__":
    main()