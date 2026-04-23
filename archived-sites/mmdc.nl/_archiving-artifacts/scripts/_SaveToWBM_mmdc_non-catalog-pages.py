#!/usr/bin/env python
"""
MMDC.nl Full WBM Re-Archive Script (Non-CATALOG_RECORDS only)

FORCE RE-ARCHIVE VERSION - Run this yourself, it will:
1. Read URLs from mmdc-urls-UNIFIED.xlsx (Sheet1)
2. Filter to only process NON-CATALOG_RECORDS URLs (467 URLs)
3. FORCE SUBMIT ALL URLs to WBM using SPN2 API (no CDX check!)
4. Update Excel columns: WBM_URL, WBM_Timestamp, Archive_Status, WBM_HTTP_Status
5. Also update local path columns if local files exist
6. Rebuild archiving-results.json with complete data

NOTE: This script forces a fresh snapshot of EVERY URL, ignoring existing archives.

ROBUSTNESS FEATURES:
- Resume capability: saves progress after EVERY URL
- Graceful shutdown: Ctrl+C saves state before exit
- Automatic retry with exponential backoff
- Excel file locking detection and retry
- Connection error recovery
- Rate limit handling with smart waits

Excel columns:
- Col 1: URL
- Col 2: Section
- Col 3: Title
- Col 4: HTTP_Status
- Col 5: Notes
- Col 6: WBM_URL
- Col 7: WBM_Timestamp
- Col 8: Archive_Status
- Col 9: WBM_HTTP_Status
- Col 10: Local_Path_Absolute
- Col 11: Local_Path_Relative
- Col 12: Local_File_Exists

USAGE:
    python SaveToWBM_mmdc_rerun.py

To reset and start fresh, delete:
    _archiving-artifacts/data/rerun-progress.json
"""

import os
import sys
import time
import json
import signal
import requests
import random
import traceback
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse, unquote

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# Configuration
SCRIPT_DIR = Path(__file__).parent
ARTIFACTS_DIR = SCRIPT_DIR.parent
DATA_DIR = ARTIFACTS_DIR / "data"
LOGS_DIR = ARTIFACTS_DIR / "logs"
MMDC_DIR = ARTIFACTS_DIR.parent  # mmdc.nl folder

# Input/Output files
EXCEL_FILE = MMDC_DIR / "mmdc-urls-UNIFIED.xlsx"
RESULTS_FILE = DATA_DIR / "archiving-results.json"
PROGRESS_FILE = DATA_DIR / "rerun-progress.json"
LOG_FILE = LOGS_DIR / "mmdc-rerun-log.txt"

# Credentials
ENV_FILE = SCRIPT_DIR / ".env"

# Rate limiting
BASE_DELAY = 1.5  # CDX queries don't need long delays
SPN_DELAY = 17  # For actual archiving submissions
MAX_RETRIES = 5  # Max retries for network errors
EXCEL_SAVE_RETRY = 5  # Times to retry Excel save
RATE_LIMIT_WAIT = 300  # 5 minutes on rate limit

# Color fills
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Global state for graceful shutdown
SHUTDOWN_REQUESTED = False
CURRENT_WORKBOOK = None
CURRENT_PROGRESS = None
RESULTS_DATA = None


def signal_handler(signum, frame):
    """Handle Ctrl+C gracefully - save state before exit."""
    global SHUTDOWN_REQUESTED
    print("\n\n" + "!" * 70)
    print("SHUTDOWN REQUESTED - Saving state...")
    print("!" * 70)
    SHUTDOWN_REQUESTED = True

    # Try to save current state
    try:
        if CURRENT_PROGRESS:
            save_progress(CURRENT_PROGRESS)
            print("Progress saved to JSON")
        if CURRENT_WORKBOOK:
            save_excel_safely(CURRENT_WORKBOOK, EXCEL_FILE)
            print("Excel file saved")
        if RESULTS_DATA:
            save_results(RESULTS_DATA)
            print("Results JSON saved")
    except Exception as e:
        print(f"Warning: Could not save all state: {e}")

    print("\nYou can resume by running the script again.")
    print("To start fresh, delete: _archiving-artifacts/data/rerun-progress.json")
    sys.exit(0)


# Register signal handler
signal.signal(signal.SIGINT, signal_handler)
if hasattr(signal, 'SIGTERM'):
    signal.signal(signal.SIGTERM, signal_handler)

# Local paths base directories (in local-archive subfolder)
LOCAL_ARCHIVE_DIR = ARTIFACTS_DIR / 'local-archive'
LOCAL_PATHS = {
    'catalog-pages': LOCAL_ARCHIVE_DIR / 'catalog-pages',
    'static-pages': LOCAL_ARCHIVE_DIR / 'static-pages',
    'images': LOCAL_ARCHIVE_DIR / 'images',
    'pdfs': LOCAL_ARCHIVE_DIR / 'pdfs',
}


def load_credentials():
    """Load IA S3 credentials from .env file."""
    if not ENV_FILE.exists():
        print(f"Warning: .env file not found at {ENV_FILE}")
        return None, None

    creds = {}
    with open(ENV_FILE, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                creds[key.strip()] = value.strip().strip("'\"")

    return creds.get('IA_ACCESS_KEY'), creds.get('IA_SECRET_KEY')


def load_progress():
    """Load progress for resume capability."""
    if PROGRESS_FILE.exists():
        try:
            with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {
        'last_row': 1,
        'total_checked': 0,
        'archived_in_wbm': 0,
        'not_in_wbm': 0,
        'newly_submitted': 0,
        'failed': 0,
        'start_time': None
    }


def save_progress(progress):
    """Save progress with error handling."""
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
            json.dump(progress, f, indent=2)
    except Exception as e:
        print(f"  Warning: Could not save progress: {e}")


def save_results(results):
    """Save results JSON."""
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(RESULTS_FILE, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"  Warning: Could not save results: {e}")


def save_excel_safely(wb, filepath, max_retries=EXCEL_SAVE_RETRY):
    """Save Excel with retry logic for file locking."""
    for attempt in range(max_retries):
        try:
            wb.save(filepath)
            return True
        except PermissionError:
            if attempt < max_retries - 1:
                wait = 5 * (attempt + 1)
                print(f"\n  Excel file locked, waiting {wait}s... (close Excel if open)")
                time.sleep(wait)
            else:
                print(f"\n  ERROR: Could not save Excel after {max_retries} attempts")
                print(f"  Please close the file and run the script again")
                return False
        except Exception as e:
            print(f"\n  Error saving Excel: {e}")
            return False
    return False


def query_cdx(url, retries=MAX_RETRIES):
    """Query CDX API to check if URL is in Wayback Machine."""
    cdx_url = f"https://web.archive.org/cdx/search/cdx?url={url}&output=json&limit=1&sort=reverse"

    for attempt in range(retries):
        try:
            response = requests.get(cdx_url, timeout=60)
            if response.status_code == 200:
                data = response.json()
                if len(data) > 1:
                    # Data format: [header, [urlkey, timestamp, original, mimetype, statuscode, digest, length]]
                    entry = data[1]
                    return {
                        'archived': True,
                        'timestamp': entry[1],
                        'http_status': entry[4],
                        'mimetype': entry[3],
                        'wbm_url': f"https://web.archive.org/web/{entry[1]}/{url}"
                    }
                else:
                    return {'archived': False}
            elif response.status_code == 429:
                wait = 30 * (attempt + 1) + random.uniform(10, 30)
                print(f"\n  CDX rate limited, waiting {wait:.0f}s...", end="", flush=True)
                time.sleep(wait)
                continue
            elif response.status_code >= 500:
                # Server error - retry
                wait = 10 * (attempt + 1)
                print(f"\n  CDX server error {response.status_code}, retry in {wait}s...", end="", flush=True)
                time.sleep(wait)
                continue
            else:
                return {'archived': False, 'error': f'HTTP {response.status_code}'}
        except requests.exceptions.Timeout:
            if attempt < retries - 1:
                wait = 10 * (attempt + 1)
                print(f"\n  CDX timeout, retry in {wait}s...", end="", flush=True)
                time.sleep(wait)
            else:
                return {'archived': False, 'error': 'Timeout after retries'}
        except requests.exceptions.ConnectionError as e:
            if attempt < retries - 1:
                wait = 15 * (attempt + 1)
                print(f"\n  Connection error, retry in {wait}s...", end="", flush=True)
                time.sleep(wait)
            else:
                return {'archived': False, 'error': f'Connection error: {str(e)[:50]}'}
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(5 * (attempt + 1))
            else:
                return {'archived': False, 'error': str(e)[:100]}

    return {'archived': False, 'error': 'Max retries exceeded'}


def submit_to_wbm(url, access_key, secret_key, retries=MAX_RETRIES):
    """Submit URL to Wayback Machine SPN2 API with retry logic."""
    if not access_key or not secret_key:
        return {'success': False, 'error': 'No credentials'}

    headers = {
        'Accept': 'application/json',
        'Authorization': f'LOW {access_key}:{secret_key}',
        'User-Agent': 'KB-Archiver/1.0 (mmdc.nl archiving project; Contact: kb.nl)'
    }

    for attempt in range(retries):
        try:
            response = requests.post(
                'https://web.archive.org/save',
                headers=headers,
                data={'url': url},
                timeout=180  # Longer timeout for actual archiving
            )

            if response.status_code == 200:
                result = response.json()
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                return {
                    'success': True,
                    'job_id': result.get('job_id', ''),
                    'wbm_url': f"https://web.archive.org/web/{timestamp}/{url}",
                    'timestamp': timestamp,
                    'message': result.get('message', '')
                }
            elif response.status_code == 429:
                if attempt < retries - 1:
                    wait = RATE_LIMIT_WAIT + random.uniform(30, 120)
                    print(f"\n  SPN RATE LIMITED! Waiting {wait:.0f}s...", end="", flush=True)
                    time.sleep(wait)
                    continue
                return {'success': False, 'error': 'Rate limited after retries', 'rate_limited': True}
            elif response.status_code == 523:
                # Origin unreachable - site may be down, don't retry
                return {'success': False, 'error': 'HTTP 523 Origin Unreachable (site down?)'}
            elif response.status_code >= 500:
                if attempt < retries - 1:
                    wait = 30 * (attempt + 1)
                    print(f"\n  SPN server error {response.status_code}, retry in {wait}s...", end="", flush=True)
                    time.sleep(wait)
                    continue
                return {'success': False, 'error': f'HTTP {response.status_code} after retries'}
            else:
                return {'success': False, 'error': f'HTTP {response.status_code}'}
        except requests.exceptions.Timeout:
            if attempt < retries - 1:
                wait = 30 * (attempt + 1)
                print(f"\n  SPN timeout, retry in {wait}s...", end="", flush=True)
                time.sleep(wait)
            else:
                return {'success': False, 'error': 'Timeout after retries'}
        except requests.exceptions.ConnectionError as e:
            if attempt < retries - 1:
                wait = 30 * (attempt + 1)
                print(f"\n  SPN connection error, retry in {wait}s...", end="", flush=True)
                time.sleep(wait)
            else:
                return {'success': False, 'error': f'Connection error: {str(e)[:50]}'}
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(15 * (attempt + 1))
            else:
                return {'success': False, 'error': str(e)[:100]}

    return {'success': False, 'error': 'Max retries exceeded'}


def find_local_path(url, section):
    """Find local file path for a URL based on its section type.

    Local archive structure:
    local-archive/
        catalog-pages/   - werk_*.html files
        static-pages/    - static HTML pages (path with __ separators)
        images/          - static asset images (with numeric prefix)
        pdfs/            - PDF documents (with numeric prefix)

    Naming conventions:
    - PDFs: /static/media/1/35/filename.pdf -> 1__35__filename.pdf
    - Images: /static/media/1/101/image.jpg -> 1__101__image.jpg
    - Static pages: /static/site/about/page -> static__site__about__page.html
    """
    parsed = urlparse(url)
    path = unquote(parsed.path)

    # PDF documents - pattern: /static/media/{num1}/{num2}/filename.pdf
    if section == 'PDF_DOCUMENTS':
        local_dir = LOCAL_PATHS.get('pdfs')
        if local_dir and local_dir.exists():
            # Parse path like: /static/media/1/35/KB_MANUSCRIPTSDATES-VOL1-INDEX.pdf
            parts = path.strip('/').split('/')
            # Expected: ['static', 'media', '1', '35', 'filename.pdf']
            if len(parts) >= 5 and parts[0] == 'static' and parts[1] == 'media':
                num1 = parts[2]
                num2 = parts[3]
                filename = parts[4]
                prefixed_name = f"{num1}__{num2}__{filename}"
                local_file = local_dir / prefixed_name
                if local_file.exists():
                    return str(local_file), f"_archiving-artifacts/local-archive/pdfs/{prefixed_name}"

            # Fallback: try just the filename (in case some files don't have prefix)
            filename = path.split('/')[-1]
            local_file = local_dir / filename
            if local_file.exists():
                return str(local_file), f"_archiving-artifacts/local-archive/pdfs/{filename}"

    # Static assets (images) - pattern: /static/media/{num1}/{num2}/image.ext
    elif section == 'STATIC_ASSETS':
        local_dir = LOCAL_PATHS.get('images')
        if local_dir and local_dir.exists():
            # Parse path like: /static/media/1/101/PA_Half-uncial_UBL_VLO_88A_f002r_detail.jpg
            parts = path.strip('/').split('/')
            if len(parts) >= 5 and parts[0] == 'static' and parts[1] == 'media':
                num1 = parts[2]
                num2 = parts[3]
                filename = parts[4]
                prefixed_name = f"{num1}__{num2}__{filename}"
                local_file = local_dir / prefixed_name
                if local_file.exists():
                    return str(local_file), f"_archiving-artifacts/local-archive/images/{prefixed_name}"

            # Fallback: try just the filename
            filename = path.split('/')[-1]
            local_file = local_dir / filename
            if local_file.exists():
                return str(local_file), f"_archiving-artifacts/local-archive/images/{filename}"

    # Catalog records
    elif section == 'CATALOG_RECORDS':
        # URL like: https://mmdc.nl/catalogus/werk/1234
        parts = path.strip('/').split('/')
        if len(parts) >= 3 and parts[0] == 'catalogus' and parts[1] == 'werk':
            werk_id = parts[2]
            local_dir = LOCAL_PATHS.get('catalog-pages')
            if local_dir and local_dir.exists():
                local_file = local_dir / f"werk_{werk_id}.html"
                if local_file.exists():
                    return str(local_file), f"_archiving-artifacts/local-archive/catalog-pages/werk_{werk_id}.html"

    # All other static pages - naming: path with __ separators + .html
    else:
        local_dir = LOCAL_PATHS.get('static-pages')
        if local_dir and local_dir.exists():
            # Convert path to double-underscore format
            # /static/site/about/AccessibilityStatement -> static__site__about__AccessibilityStatement.html
            clean_path = path.strip('/').replace('/', '__')

            # Try with .html extension
            for pattern in [f"{clean_path}.html", clean_path, f"{clean_path}__index.html"]:
                local_file = local_dir / pattern
                if local_file.exists():
                    return str(local_file), f"_archiving-artifacts/local-archive/static-pages/{pattern}"

            # Also try URL-encoded versions (some filenames have %20 etc)
            from urllib.parse import quote
            encoded_path = path.strip('/').replace('/', '__')
            encoded_path = quote(encoded_path, safe='_')
            for pattern in [f"{encoded_path}.html", encoded_path]:
                local_file = local_dir / pattern
                if local_file.exists():
                    return str(local_file), f"_archiving-artifacts/local-archive/static-pages/{pattern}"

            # Search for any file containing key parts as fallback
            if path:
                key_parts = [p for p in path.strip('/').split('/') if p and len(p) > 3]
                if key_parts:
                    try:
                        for f in local_dir.iterdir():
                            if f.is_file():
                                fname = f.name.lower()
                                # Check if the last 2 parts of the path are in the filename
                                if all(part.lower() in fname for part in key_parts[-2:]):
                                    return str(f), f"_archiving-artifacts/local-archive/static-pages/{f.name}"
                    except:
                        pass

    return None, None


def main():
    global CURRENT_WORKBOOK, CURRENT_PROGRESS, RESULTS_DATA, SHUTDOWN_REQUESTED

    print("=" * 70)
    print("MMDC.nl WBM Re-Archive (Non-CATALOG_RECORDS)")
    print("=" * 70)
    print()
    print("ROBUSTNESS: Ctrl+C saves state | Auto-retry on errors | Resume capable")
    print("=" * 70)
    print()

    # Ensure directories exist
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)

    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    # Load credentials for archiving
    access_key, secret_key = load_credentials()
    if access_key:
        print(f"Credentials loaded: {access_key[:4]}...{access_key[-4:]}")
        print("Will submit unarchived URLs to Wayback Machine")
    else:
        print("WARNING: No credentials - will only check CDX status, not submit new URLs")
        print("To submit URLs, create .env file with IA_ACCESS_KEY and IA_SECRET_KEY")

    # Load Excel
    print(f"\nLoading {EXCEL_FILE}...")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        CURRENT_WORKBOOK = wb
    except Exception as e:
        print(f"ERROR: Could not load Excel: {e}")
        sys.exit(1)

    ws = wb['Sheet1']

    # Collect non-CATALOG_RECORDS rows
    rows_to_process = []
    for row in range(2, ws.max_row + 1):
        section = ws.cell(row=row, column=2).value
        if section != 'CATALOG_RECORDS':
            url = ws.cell(row=row, column=1).value
            if url:
                rows_to_process.append({
                    'row': row,
                    'url': url,
                    'section': section,
                    'title': ws.cell(row=row, column=3).value
                })

    print(f"Found {len(rows_to_process)} non-CATALOG_RECORDS URLs to process")

    # Section breakdown
    sections = {}
    for item in rows_to_process:
        sections[item['section']] = sections.get(item['section'], 0) + 1
    print("\nBy section:")
    for sec, cnt in sorted(sections.items(), key=lambda x: -x[1]):
        print(f"  {sec}: {cnt}")

    # Load progress
    progress = load_progress()
    CURRENT_PROGRESS = progress

    if progress['start_time'] and progress['last_row'] > 1:
        already_done = sum(1 for item in rows_to_process if item['row'] <= progress['last_row'])
        print(f"\nRESUMING from previous run:")
        print(f"  Started: {progress['start_time']}")
        print(f"  Last row: {progress['last_row']}")
        print(f"  Already processed: ~{already_done}")
        print(f"  Remaining: ~{len(rows_to_process) - already_done}")
    else:
        progress['start_time'] = datetime.now().isoformat()

    # Results for JSON
    results = {
        'generated_at': datetime.now().isoformat(),
        'excel_file': str(EXCEL_FILE),
        'total_urls': len(rows_to_process),
        'urls': []
    }
    RESULTS_DATA = results

    # Estimate time - FORCE RE-ARCHIVE: ALL URLs will be submitted
    remaining = sum(1 for item in rows_to_process if item['row'] > progress['last_row'])
    if access_key:
        est_time = remaining * (SPN_DELAY + 5) / 60
        print(f"\n** FORCE RE-ARCHIVE MODE **")
        print(f"Estimated time: ~{est_time:.0f} minutes for {remaining} URLs (ALL will be submitted!)")
    else:
        print(f"\n** NO CREDENTIALS - cannot submit URLs! **")
        print(f"Please create .env file with IA_ACCESS_KEY and IA_SECRET_KEY")

    print("\nStarting in 3 seconds... (Ctrl+C to abort)")
    time.sleep(3)

    print("\n" + "-" * 70)
    print("Processing URLs...")
    print("-" * 70 + "\n")

    # Open log file
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as log:
        log.write(f"\n{'='*70}\n")
        log.write(f"MMDC.nl Re-Archive Run: {datetime.now().isoformat()}\n")
        log.write(f"{'='*70}\n\n")

        # Process each URL
        for idx, item in enumerate(rows_to_process):
            # Check for shutdown request
            if SHUTDOWN_REQUESTED:
                break

            row = item['row']
            url = item['url']
            section = item['section']

            # Skip already processed rows
            if row <= progress['last_row'] and progress['last_row'] > 1:
                # Load existing result into results array
                existing_status = ws.cell(row=row, column=8).value
                if existing_status:
                    results['urls'].append({
                        'url': url,
                        'section': section,
                        'title': item['title'],
                        'row': row,
                        'status': existing_status,
                        'wbm_url': ws.cell(row=row, column=6).value,
                        'wbm_timestamp': ws.cell(row=row, column=7).value,
                        'skipped_resume': True
                    })
                continue

            progress['total_checked'] += 1
            pct = (progress['total_checked'] * 100) // len(rows_to_process)

            url_short = url[-55:] if len(url) > 55 else url
            print(f"[{progress['total_checked']}/{len(rows_to_process)}] ({pct}%) {section}: ...{url_short}", end=" ", flush=True)

            result_entry = {
                'url': url,
                'section': section,
                'title': item['title'],
                'row': row,
                'checked_at': datetime.now().isoformat()
            }

            # FORCE RE-ARCHIVE: Submit ALL URLs to WBM (no CDX check!)
            if access_key and secret_key:
                print("SUBMITTING...", end=" ", flush=True)
                submit_result = submit_to_wbm(url, access_key, secret_key)

                if submit_result.get('success'):
                    job_id = submit_result.get('job_id', 'N/A')
                    print(f"SUBMITTED (job: {job_id[:20] if len(job_id) > 20 else job_id})")
                    ws.cell(row=row, column=6).value = submit_result['wbm_url']
                    ws.cell(row=row, column=7).value = submit_result['timestamp']
                    ws.cell(row=row, column=8).value = 'SUBMITTED'
                    ws.cell(row=row, column=8).fill = YELLOW_FILL
                    progress['newly_submitted'] += 1
                    result_entry['status'] = 'SUBMITTED'
                    result_entry['job_id'] = job_id
                    result_entry['wbm_url'] = submit_result['wbm_url']

                    # Rate limit delay for SPN
                    delay = SPN_DELAY + random.uniform(2, 8)
                    time.sleep(delay)
                else:
                    error = submit_result.get('error', 'Unknown')
                    print(f"FAILED: {error}")
                    ws.cell(row=row, column=8).value = f'FAILED: {error[:30]}'
                    ws.cell(row=row, column=8).fill = RED_FILL
                    progress['failed'] += 1
                    result_entry['status'] = 'FAILED'
                    result_entry['error'] = error

                    # Extra delay on failure
                    time.sleep(5)
            else:
                print("NO CREDENTIALS")
                ws.cell(row=row, column=8).value = 'NO_CREDENTIALS'
                ws.cell(row=row, column=8).fill = RED_FILL
                progress['not_in_wbm'] += 1
                result_entry['status'] = 'NO_CREDENTIALS'

            # Check for local path
            abs_path, rel_path = find_local_path(url, section)
            if abs_path:
                ws.cell(row=row, column=10).value = abs_path
                ws.cell(row=row, column=11).value = rel_path
                ws.cell(row=row, column=12).value = 'TRUE'
                result_entry['local_path'] = rel_path
                result_entry['local_exists'] = True
            else:
                ws.cell(row=row, column=12).value = 'FALSE'
                result_entry['local_exists'] = False

            results['urls'].append(result_entry)

            # Update progress
            progress['last_row'] = row
            save_progress(progress)

            # Log
            try:
                log.write(f"{result_entry.get('status', 'UNKNOWN')}: {url}\n")
                log.flush()
            except:
                pass

            # Save Excel and JSON every 5 URLs (streaming update)
            if progress['total_checked'] % 5 == 0:
                if save_excel_safely(wb, EXCEL_FILE):
                    print(f"  [Excel saved at {progress['total_checked']} | {progress['newly_submitted']} submitted, {progress['failed']} failed]")
                # Also save results JSON
                results['summary'] = {
                    'total_checked': progress['total_checked'],
                    'newly_submitted': progress['newly_submitted'],
                    'failed': progress['failed'],
                    'last_updated': datetime.now().isoformat(),
                    'status': 'in_progress'
                }
                save_results(results)

    # Final save
    print("\n" + "-" * 70)
    print("Saving final state...")
    save_excel_safely(wb, EXCEL_FILE)

    # Update results summary
    results['summary'] = {
        'total_checked': progress['total_checked'],
        'newly_submitted': progress['newly_submitted'],
        'failed': progress['failed'],
        'completed_at': datetime.now().isoformat(),
        'mode': 'FORCE_RE_ARCHIVE'
    }

    # Save results JSON
    save_results(results)

    # Final progress save
    progress['end_time'] = datetime.now().isoformat()
    save_progress(progress)

    print()
    print("=" * 70)
    print("COMPLETE" if not SHUTDOWN_REQUESTED else "INTERRUPTED (state saved)")
    print("=" * 70)
    print(f"Total processed: {progress['total_checked']}")
    print(f"Submitted to WBM: {progress['newly_submitted']}")
    print(f"Failed: {progress['failed']}")
    print()
    print(f"Excel updated: {EXCEL_FILE}")
    print(f"Results JSON: {RESULTS_FILE}")
    print(f"Log file: {LOG_FILE}")

    if not SHUTDOWN_REQUESTED and progress['total_checked'] == len(rows_to_process):
        # Clean up progress file on complete success
        print("\nAll URLs processed! You can delete the progress file to reset.")
    else:
        print(f"\nTo resume, run the script again.")
        print(f"To start fresh, delete: {PROGRESS_FILE}")


if __name__ == "__main__":
    main()
