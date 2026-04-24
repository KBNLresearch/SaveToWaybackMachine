#!/usr/bin/env python
"""
Save manuscripts.kb.nl Wiki Priority URLs to Wayback Machine.

Phase 1 of the manuscripts.kb.nl preservation project (Dec 2025).
Archives URLs that are linked from Dutch Wikipedia and Wikimedia Commons
first, because broken Wikimedia links are the most visible impact of
the site shutdown.

Workflow:
  1. Query the MediaWiki exturlusage API on nl.wikipedia.org and
     commons.wikimedia.org to discover all pages linking to manuscripts.kb.nl
  2. Deduplicate URLs (some appear on both wikis)
  3. Submit each URL to the Internet Archive's SPN2 API
  4. Write results to wiki-priority-urls-WBM.xlsx (streaming, every 10 URLs)

Resume capability:
  - Progress JSON saved after every URL (archiving-progress.json)
  - Can resume from exact index after interruption

Output files:
  - wiki-priority-urls-WBM.xlsx: Excel with URL, source wiki, WBM status
  - wpnl_priority_urls.json: raw Wikipedia API response
  - commons_priority_urls.json: raw Commons API response
  - wbm-archiving-results.json: full progress + results

Requirements:
  - .env file with IA_ACCESS_KEY and IA_SECRET_KEY
  - openpyxl, requests

Usage:
  python SaveToWBM_manuscripts_wiki_priority.py
"""

import os
import sys
import time
import json
import requests
import random
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export disabled.")
    print("Install with: pip install openpyxl")

# User-Agent required by Wikimedia APIs
USER_AGENT = "KB-Archiver/1.0 (https://www.kb.nl/; manuscripts.kb.nl archiving project) Python/requests"

# Configuration
SCRIPT_DIR = Path(__file__).parent
ARTIFACTS_DIR = SCRIPT_DIR.parent
DATA_DIR = ARTIFACTS_DIR / "data" / "wikiPriorityArchiving"
LOGS_DIR = ARTIFACTS_DIR / "logs"

# Output files
WPNL_URLS_FILE = DATA_DIR / "wpnl_priority_urls.json"
COMMONS_URLS_FILE = DATA_DIR / "commons_priority_urls.json"
OUTPUT_FILE = DATA_DIR / "wbm-archiving-results.json"
LOG_FILE = LOGS_DIR / "wiki-priority-archiving-log.txt"
PROGRESS_FILE = DATA_DIR / "archiving-progress.json"
EXCEL_FILE = ARTIFACTS_DIR.parent / "wiki-priority-urls-WBM.xlsx"

# Rate limiting - 17s is safe for authenticated users (15 req/min limit = 4s min)
BASE_DELAY = 17  # seconds between requests
MAX_RETRIES = 3
BACKOFF_MULTIPLIER = 2
RATE_LIMIT_PAUSE = 300  # 5 minutes pause on rate limit

# Load credentials from .env file (same folder as this script)
ENV_FILE = SCRIPT_DIR / ".env"


def load_credentials():
    """Load IA S3 credentials from .env file."""
    if not ENV_FILE.exists():
        raise FileNotFoundError(f".env file not found at {ENV_FILE}")

    creds = {}
    with open(ENV_FILE, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                creds[key.strip()] = value.strip().strip("'\"")

    access_key = creds.get('IA_ACCESS_KEY')
    secret_key = creds.get('IA_SECRET_KEY')

    if not access_key or not secret_key:
        raise ValueError("IA_ACCESS_KEY and IA_SECRET_KEY must be set in .env file")

    return access_key, secret_key


def fetch_wikipedia_urls():
    """Fetch manuscripts.kb.nl URLs linked from Dutch Wikipedia."""
    print("Fetching URLs from Dutch Wikipedia...")

    api_url = "https://nl.wikipedia.org/w/api.php"
    all_urls = []
    continue_token = None

    while True:
        params = {
            "action": "query",
            "list": "exturlusage",
            "euquery": "manuscripts.kb.nl",
            "eulimit": "500",
            "eunamespace": "0",  # Main namespace only
            "format": "json"
        }
        if continue_token:
            params["eucontinue"] = continue_token

        try:
            headers = {"User-Agent": USER_AGENT}
            response = requests.get(api_url, params=params, headers=headers, timeout=30)
            data = response.json()

            if "query" in data and "exturlusage" in data["query"]:
                for item in data["query"]["exturlusage"]:
                    url = item.get("url", "")
                    if url and "manuscripts.kb.nl" in url:
                        all_urls.append({
                            "url": url,
                            "source": "Wikipedia NL",
                            "page_title": item.get("title", ""),
                            "page_id": item.get("pageid", "")
                        })

            # Check for continuation
            if "continue" in data and "eucontinue" in data["continue"]:
                continue_token = data["continue"]["eucontinue"]
                print(f"  Found {len(all_urls)} URLs so far, continuing...")
            else:
                break

        except Exception as e:
            print(f"  Error fetching from Wikipedia: {e}")
            break

    print(f"  Total Wikipedia URLs: {len(all_urls)}")
    return all_urls


def fetch_commons_urls():
    """Fetch manuscripts.kb.nl URLs linked from Wikimedia Commons."""
    print("Fetching URLs from Wikimedia Commons...")

    api_url = "https://commons.wikimedia.org/w/api.php"
    all_urls = []
    continue_token = None

    while True:
        params = {
            "action": "query",
            "list": "exturlusage",
            "euquery": "manuscripts.kb.nl",
            "eulimit": "500",
            "eunamespace": "6",  # File namespace
            "format": "json"
        }
        if continue_token:
            params["eucontinue"] = continue_token

        try:
            headers = {"User-Agent": USER_AGENT}
            response = requests.get(api_url, params=params, headers=headers, timeout=30)
            data = response.json()

            if "query" in data and "exturlusage" in data["query"]:
                for item in data["query"]["exturlusage"]:
                    url = item.get("url", "")
                    if url and "manuscripts.kb.nl" in url:
                        all_urls.append({
                            "url": url,
                            "source": "Commons",
                            "page_title": item.get("title", ""),
                            "page_id": item.get("pageid", "")
                        })

            # Check for continuation
            if "continue" in data and "eucontinue" in data["continue"]:
                continue_token = data["continue"]["eucontinue"]
                print(f"  Found {len(all_urls)} URLs so far, continuing...")
            else:
                break

        except Exception as e:
            print(f"  Error fetching from Commons: {e}")
            break

    print(f"  Total Commons URLs: {len(all_urls)}")
    return all_urls


def deduplicate_urls(wiki_urls, commons_urls):
    """Merge and deduplicate URLs, keeping source info."""
    seen = {}

    # Wiki URLs first (higher priority)
    for item in wiki_urls:
        url = item["url"]
        if url not in seen:
            seen[url] = item
        else:
            seen[url]["also_on"] = item["source"]

    # Commons URLs
    for item in commons_urls:
        url = item["url"]
        if url not in seen:
            seen[url] = item
        else:
            seen[url]["also_on"] = item["source"]

    return list(seen.values())


def load_progress():
    """Load progress from previous run (for resume capability)."""
    if PROGRESS_FILE.exists():
        try:
            with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"  Warning: Could not load progress file: {e}")

    return {
        'start_time': None,
        'end_time': None,
        'duration_seconds': 0,
        'current_index': 0,
        'total_urls': 0,
        'successful': 0,
        'failed': 0,
        'skipped': 0,
        'last_url': None,
        'last_job_id': None,
        'results': []
    }


def save_progress(progress, start_time=None):
    """Save progress for resume capability."""
    if start_time:
        progress['duration_seconds'] = (datetime.now() - start_time).total_seconds()
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, indent=2)


def save_to_wayback(url, access_key, secret_key, retries=0):
    """Submit a single URL to the Wayback Machine via the SPN2 API.

    Returns a dict with 'success' (bool), and on success: 'job_id', 'wbm_url',
    'message'. On failure: 'error', and optionally 'rate_limited' (bool).
    """
    headers = {
        'Accept': 'application/json',
        'Authorization': f'LOW {access_key}:{secret_key}',
        'User-Agent': USER_AGENT
    }

    data = {
        'url': url,
    }

    try:
        response = requests.post(
            'https://web.archive.org/save',
            headers=headers,
            data=data,
            timeout=120
        )

        if response.status_code == 200:
            result = response.json()
            job_id = result.get('job_id', '')
            message = result.get('message', 'Submitted')

            # Construct WBM URL
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            wbm_url = f"https://web.archive.org/web/{timestamp}/{url}"

            return {
                'success': True,
                'job_id': job_id,
                'wbm_url': wbm_url,
                'message': message,
                'url': url,
                'retries': retries
            }
        elif response.status_code == 429:
            return {
                'success': False,
                'error': 'HTTP 429 Rate Limited',
                'rate_limited': True,
                'url': url,
                'retries': retries
            }
        elif response.status_code == 523:
            return {
                'success': False,
                'error': f'HTTP 523 Origin Unreachable (site may be down)',
                'url': url,
                'retries': retries
            }
        else:
            return {
                'success': False,
                'error': f'HTTP {response.status_code}',
                'response': response.text[:200] if response.text else '',
                'url': url,
                'retries': retries
            }
    except requests.exceptions.Timeout:
        return {
            'success': False,
            'error': 'Timeout (120s)',
            'url': url,
            'retries': retries
        }
    except requests.exceptions.ConnectionError as e:
        return {
            'success': False,
            'error': f'Connection Error: {str(e)[:100]}',
            'url': url,
            'retries': retries
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)[:200],
            'url': url,
            'retries': retries
        }


def update_excel_row(wb, row, result):
    """Update a row in the Excel file with WBM results."""
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    if result.get('success'):
        ws.cell(row=row, column=5, value='OK')
        ws.cell(row=row, column=5).fill = green_fill
        ws.cell(row=row, column=6, value=result.get('job_id', ''))
        ws.cell(row=row, column=7, value=result.get('wbm_url', ''))
        ws.cell(row=row, column=8, value=result.get('timestamp', ''))
    else:
        ws.cell(row=row, column=5, value='FAILED')
        ws.cell(row=row, column=5).fill = red_fill
        ws.cell(row=row, column=9, value=result.get('error', ''))
        ws.cell(row=row, column=8, value=result.get('timestamp', ''))


def setup_excel(all_urls):
    """Create Excel file with URLs and headers."""
    if not EXCEL_AVAILABLE:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wiki Priority URLs"

    # Headers
    headers = [
        ("URL", 60),
        ("Source", 15),
        ("Wiki Page Title", 35),
        ("Wiki Page ID", 12),
        ("WBM Status", 12),
        ("WBM Job ID", 45),
        ("WBM URL", 70),
        ("Archived At", 22),
        ("Error", 30)
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Add URL data
    for row_idx, item in enumerate(all_urls, 2):
        ws.cell(row=row_idx, column=1, value=item.get('url', ''))
        ws.cell(row=row_idx, column=2, value=item.get('source', ''))
        ws.cell(row=row_idx, column=3, value=item.get('page_title', ''))
        ws.cell(row=row_idx, column=4, value=item.get('page_id', ''))

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(EXCEL_FILE)
    return wb


def main():
    print("=" * 70)
    print("manuscripts.kb.nl Wiki Priority Archiver")
    print("Wayback Machine SPN2 API (Authenticated)")
    print("=" * 70)
    print()

    # Ensure directories exist
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)

    # Load credentials
    print("Loading credentials...")
    try:
        access_key, secret_key = load_credentials()
        print(f"  Access key: {access_key[:4]}...{access_key[-4:]}")
        print("  Credentials loaded successfully")
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    # Step 1: Fetch Wiki URLs
    print()
    print("-" * 70)
    print("STEP 1: Fetching Wiki Priority URLs")
    print("-" * 70)

    wiki_urls = fetch_wikipedia_urls()
    commons_urls = fetch_commons_urls()

    # Save raw fetched URLs
    with open(WPNL_URLS_FILE, 'w', encoding='utf-8') as f:
        json.dump({"fetched": datetime.now().isoformat(), "urls": wiki_urls}, f, indent=2)
    print(f"  Saved Wikipedia URLs to {WPNL_URLS_FILE}")

    with open(COMMONS_URLS_FILE, 'w', encoding='utf-8') as f:
        json.dump({"fetched": datetime.now().isoformat(), "urls": commons_urls}, f, indent=2)
    print(f"  Saved Commons URLs to {COMMONS_URLS_FILE}")

    # Deduplicate
    all_priority_urls = deduplicate_urls(wiki_urls, commons_urls)
    print(f"\n  Total unique priority URLs: {len(all_priority_urls)}")

    if not all_priority_urls:
        print("\nNo URLs found to archive!")
        sys.exit(0)

    # Setup Excel
    print(f"\n  Setting up Excel: {EXCEL_FILE}")
    wb = setup_excel(all_priority_urls)

    # Load progress
    progress = load_progress()

    # Check if resuming
    start_index = progress.get('current_index', 0)
    if start_index > 0 and progress.get('results'):
        print(f"\n  Resuming from previous run:")
        print(f"    Current index: {start_index}")
        print(f"    Successful: {progress.get('successful', 0)}")
        print(f"    Failed: {progress.get('failed', 0)}")

        # Restore Excel with previous results
        if wb:
            results_by_url = {r['url']: r for r in progress['results'] if 'url' in r}
            for row_idx, item in enumerate(all_priority_urls, 2):
                url = item.get('url')
                if url in results_by_url:
                    update_excel_row(wb, row_idx, results_by_url[url])
            wb.save(EXCEL_FILE)

        start_time = datetime.fromisoformat(progress['start_time']) if progress.get('start_time') else datetime.now()
    else:
        start_time = datetime.now()
        progress['start_time'] = start_time.isoformat()
        progress['total_urls'] = len(all_priority_urls)

    # Step 2: Archive to WBM
    print()
    print("-" * 70)
    print("STEP 2: Archiving to Wayback Machine")
    print("-" * 70)

    # Estimate time
    remaining = len(all_priority_urls) - start_index
    est_hours = (remaining * BASE_DELAY) / 3600
    print(f"\n  Remaining URLs: {remaining}")
    print(f"  Estimated time: ~{est_hours:.1f} hours")
    print(f"  Delay between requests: {BASE_DELAY}s")
    print()

    # Open log file
    log_mode = 'a' if start_index > 0 else 'w'
    with open(LOG_FILE, log_mode, encoding='utf-8') as log:
        if log_mode == 'w':
            log.write(f"Wiki Priority Archiving Log - manuscripts.kb.nl\n")
            log.write(f"Started: {datetime.now().isoformat()}\n")
            log.write("=" * 70 + "\n\n")
        else:
            log.write(f"\n--- Resuming at {datetime.now().isoformat()} ---\n\n")

        total_urls = len(all_priority_urls)

        for i in range(start_index, total_urls):
            item = all_priority_urls[i]
            url = item['url']
            source = item.get('source', 'unknown')
            row = i + 2  # Excel row (1-indexed + header)

            # Progress display
            pct = (i * 100) // total_urls if total_urls > 0 else 0
            url_short = url[-50:] if len(url) > 50 else url
            print(f"    [{i+1}/{total_urls}] ({pct}%) [{source[:8]}] ...{url_short}", end=" ", flush=True)

            # Submit to WBM with retry logic
            retry_count = 0
            current_delay = BASE_DELAY
            result = None

            while retry_count <= MAX_RETRIES:
                result = save_to_wayback(url, access_key, secret_key, retries=retry_count)

                if result['success']:
                    break
                elif result.get('rate_limited'):
                    if retry_count < MAX_RETRIES:
                        wait_time = RATE_LIMIT_PAUSE + random.uniform(30, 120)
                        print(f"\n    RATE LIMITED! Waiting {wait_time:.0f}s before retry...", end=" ", flush=True)
                        try:
                            log.write(f"RATE LIMITED at {datetime.now().isoformat()} - waiting {wait_time:.0f}s\n")
                            log.flush()
                        except OSError:
                            pass
                        time.sleep(wait_time)
                        retry_count += 1
                    else:
                        break
                else:
                    if retry_count < MAX_RETRIES:
                        backoff = current_delay * BACKOFF_MULTIPLIER + random.uniform(5, 15)
                        print(f"RETRY ({retry_count + 1})...", end=" ", flush=True)
                        time.sleep(backoff)
                        retry_count += 1
                        current_delay = backoff
                    else:
                        break

            # Process result
            result['timestamp'] = datetime.now().isoformat()
            result['source'] = source
            result['index'] = i

            if result['success']:
                job_id = result.get('job_id', 'no-job-id')
                message = result.get('message', '')
                print(f"OK | job: {job_id[-20:]} | {message[:50]}")
                progress['successful'] += 1
                progress['last_url'] = url
                progress['last_job_id'] = job_id
                try:
                    log.write(f"OK: {url}\n  Job: {job_id}\n  Source: {source}\n")
                    log.flush()
                except OSError:
                    pass
            else:
                error_msg = result.get('error', 'Unknown error')
                print(f"FAILED: {error_msg}")
                progress['failed'] += 1
                try:
                    log.write(f"FAILED: {url}\n  Error: {error_msg}\n  Source: {source}\n")
                    log.flush()
                except OSError:
                    pass

            # Update Excel row
            if wb:
                update_excel_row(wb, row, result)

            # Add to results
            progress['results'].append(result)
            progress['current_index'] = i + 1

            # Save progress
            save_progress(progress, start_time)

            # Save Excel every 10 successful URLs
            if wb and progress['successful'] % 10 == 0:
                try:
                    wb.save(EXCEL_FILE)
                    print(f"    [Excel saved at {progress['successful']} archived]")
                except PermissionError:
                    print(f"    WARNING: Could not save Excel (file open?)")

            print(f"    [Progress: {progress['successful']} OK, {progress['failed']} FAIL | {progress['duration_seconds']:.0f}s]")

            # Rate limiting delay
            if i < total_urls - 1:
                delay = BASE_DELAY + random.uniform(2, 8)
                time.sleep(delay)

        # Final summary
        end_time = datetime.now()
        progress['end_time'] = end_time.isoformat()
        save_progress(progress, start_time)

        try:
            log.write("\n" + "=" * 70 + "\n")
            log.write(f"ARCHIVING COMPLETE\n")
            log.write(f"Finished: {end_time.isoformat()}\n")
            log.write(f"Total archived: {progress['successful']}\n")
            log.write(f"Total failed: {progress['failed']}\n")
        except OSError:
            pass

    # Final Excel save
    if wb:
        wb.save(EXCEL_FILE)

    # Save final results JSON
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, indent=2)

    print("\n" + "=" * 70)
    print("WIKI PRIORITY ARCHIVING COMPLETE")
    print("=" * 70)
    print(f"Total archived: {progress['successful']}")
    print(f"Total failed: {progress['failed']}")
    print(f"Duration: {progress['duration_seconds']:.0f}s")
    print(f"\nExcel: {EXCEL_FILE}")
    print(f"Progress: {PROGRESS_FILE}")
    print(f"Results: {OUTPUT_FILE}")
    print(f"Log: {LOG_FILE}")

    if progress['failed'] > 0:
        print(f"\nWARNING: {progress['failed']} URLs failed to archive!")


if __name__ == "__main__":
    main()
