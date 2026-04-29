#!/usr/bin/env python
"""
Look up actual WBM capture URLs and timestamps for manuscripts.kb.nl.

For each URL in manuscripts-urls-wbm-archived.xlsx that has a submission but
no capture data yet, queries the CDX API to find the actual archived capture
closest to the submission timestamp, then writes WBM_URL_capture and
WBM_Timestamp_capture back into the Excel.

Resume-safe: saves progress after every batch of Excel writes.
"""

import os
import sys
import time
import json
import random
import requests
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
ARTIFACTS_DIR = SCRIPT_DIR.parent
PROJECT_DIR = ARTIFACTS_DIR.parent
EXCEL_FILE = PROJECT_DIR / "manuscripts-urls-wbm-archived.xlsx"
LOG_FILE = ARTIFACTS_DIR / "logs" / "capture-lookup-log.txt"

# Load .env
from dotenv import load_dotenv
load_dotenv(ARTIFACTS_DIR / ".env")

USER_AGENT = "KB-Archiver/1.0 (https://www.kb.nl/; manuscripts.kb.nl archiving project) Python/requests"
CDX_BASE = "https://web.archive.org/cdx/search/cdx"
MAX_RETRIES = 5
BASE_DELAY = 1.5  # seconds between CDX requests (CDX is read-only, less restrictive)
EXCEL_SAVE_INTERVAL = 5  # save Excel every N successful lookups

# Column indices (1-based) matching the new schema
COL_URL = 1
COL_WBM_URL_SUBMISSION = 8
COL_WBM_TIMESTAMP_SUBMISSION = 9
COL_WBM_URL_CAPTURE = 10
COL_WBM_TIMESTAMP_CAPTURE = 11


def log(msg):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def extract_timestamp_from_wbm_url(wbm_url):
    """Extract the 14-digit timestamp from a WBM URL like .../web/20251211014959/..."""
    if not wbm_url or "/web/" not in wbm_url:
        return None
    parts = wbm_url.split("/web/")
    if len(parts) < 2:
        return None
    ts_part = parts[1].split("/")[0]
    if len(ts_part) >= 14 and ts_part[:14].isdigit():
        return ts_part[:14]
    return None


def query_cdx(url, submission_timestamp=None, retries=MAX_RETRIES):
    """
    Query CDX API for the capture closest to the submission timestamp.
    If submission_timestamp is given (14-digit), find the nearest capture.
    Otherwise, find the most recent capture.
    """
    params = {
        "url": url,
        "output": "json",
        "limit": 5,
    }
    if submission_timestamp:
        params["closest"] = submission_timestamp
        params["sort"] = "closest"
    else:
        params["sort"] = "reverse"
        params["limit"] = 1

    for attempt in range(retries):
        try:
            response = requests.get(
                CDX_BASE, params=params, timeout=60,
                headers={"User-Agent": USER_AGENT}
            )
            if response.status_code == 200:
                data = response.json()
                if len(data) > 1:
                    # [header, [urlkey, timestamp, original, mimetype, statuscode, digest, length]]
                    entry = data[1]
                    ts = entry[1]
                    original_url = entry[2]
                    capture_url = f"https://web.archive.org/web/{ts}/{original_url}"
                    # Format timestamp as ISO
                    ts_iso = f"{ts[0:4]}-{ts[4:6]}-{ts[6:8]}T{ts[8:10]}:{ts[10:12]}:{ts[12:14]}"
                    return {
                        "found": True,
                        "capture_url": capture_url,
                        "capture_timestamp": ts_iso,
                        "raw_timestamp": ts,
                        "http_status": entry[4],
                    }
                else:
                    return {"found": False}
            elif response.status_code == 429:
                wait = 30 * (attempt + 1) + random.uniform(10, 30)
                log(f"  CDX rate limited, waiting {wait:.0f}s...")
                time.sleep(wait)
                continue
            elif response.status_code >= 500:
                wait = 10 * (attempt + 1)
                log(f"  CDX server error {response.status_code}, retry in {wait}s...")
                time.sleep(wait)
                continue
            else:
                return {"found": False, "error": f"HTTP {response.status_code}"}
        except requests.exceptions.Timeout:
            if attempt < retries - 1:
                wait = 10 * (attempt + 1)
                log(f"  CDX timeout, retry in {wait}s...")
                time.sleep(wait)
            else:
                return {"found": False, "error": "Timeout after retries"}
        except requests.exceptions.ConnectionError as e:
            if attempt < retries - 1:
                wait = 15 * (attempt + 1)
                log(f"  Connection error, retry in {wait}s...")
                time.sleep(wait)
            else:
                return {"found": False, "error": f"Connection error: {str(e)[:50]}"}
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(5 * (attempt + 1))
            else:
                return {"found": False, "error": str(e)[:100]}

    return {"found": False, "error": "Max retries exceeded"}


def main():
    log(f"Starting capture lookup for {EXCEL_FILE.name}")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    # Process wiki_priority first (needed for Wikipedia/Commons link updates),
    # then the remaining per-type sheets, then ALL_URLS last
    sheets_to_process = []
    if "wiki_priority" in wb.sheetnames:
        sheets_to_process.append("wiki_priority")
    for s in wb.sheetnames:
        if s not in ("ALL_URLS", "wiki_priority"):
            sheets_to_process.append(s)
    sheets_to_process.append("ALL_URLS")

    total_looked_up = 0
    total_found = 0
    total_skipped = 0
    total_not_found = 0
    unsaved_count = 0

    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]
        row_count = ws.max_row - 1
        log(f"\n--- Sheet: {sheet_name} ({row_count} rows) ---")

        sheet_looked_up = 0
        sheet_found = 0

        for row_idx in range(2, ws.max_row + 1):
            url = ws.cell(row=row_idx, column=COL_URL).value
            if not url:
                continue

            # Skip if capture already filled
            existing_capture = ws.cell(row=row_idx, column=COL_WBM_URL_CAPTURE).value
            if existing_capture:
                total_skipped += 1
                continue

            # Need a submission URL/timestamp to look up
            wbm_url_sub = ws.cell(row=row_idx, column=COL_WBM_URL_SUBMISSION).value
            if not wbm_url_sub:
                total_skipped += 1
                continue

            submission_ts = extract_timestamp_from_wbm_url(wbm_url_sub)

            # Query CDX
            result = query_cdx(url.strip(), submission_ts)
            sheet_looked_up += 1
            total_looked_up += 1

            if result.get("found"):
                ws.cell(row=row_idx, column=COL_WBM_URL_CAPTURE).value = result["capture_url"]
                ws.cell(row=row_idx, column=COL_WBM_TIMESTAMP_CAPTURE).value = result["capture_timestamp"]
                sheet_found += 1
                total_found += 1
                unsaved_count += 1
                log(f"  [{sheet_name}] {sheet_looked_up}/{row_count} | "
                    f"OK: {url[:60]}... -> {result['raw_timestamp']}")
            else:
                total_not_found += 1
                error = result.get("error", "not in CDX")
                log(f"  [{sheet_name}] {sheet_looked_up}/{row_count} | "
                    f"NOT FOUND: {url[:60]} ({error})")

            # Save Excel periodically
            if unsaved_count >= EXCEL_SAVE_INTERVAL:
                try:
                    wb.save(EXCEL_FILE)
                    log(f"  Saved Excel ({total_found} captures so far)")
                    unsaved_count = 0
                except PermissionError:
                    log(f"  WARNING: Excel file locked, will retry on next interval")

            # Rate limit
            time.sleep(BASE_DELAY + random.uniform(0.2, 0.8))

        log(f"  Sheet {sheet_name} done: {sheet_found}/{sheet_looked_up} captures found")

    # Final save
    try:
        wb.save(EXCEL_FILE)
        log(f"\nFinal save complete.")
    except PermissionError:
        log(f"\nWARNING: Could not save Excel (file locked). Close Excel and re-run.")

    log(f"\n=== SUMMARY ===")
    log(f"Total looked up: {total_looked_up}")
    log(f"Total captures found: {total_found}")
    log(f"Total not found in CDX: {total_not_found}")
    log(f"Total skipped (already done or no submission): {total_skipped}")


if __name__ == "__main__":
    main()
