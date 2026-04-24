#!/usr/bin/env python
"""
Save manuscripts.kb.nl URLs to Wayback Machine - Excel Sheet by Sheet.

Bulk archiving script for Phase 2 of the manuscripts.kb.nl preservation
project (Dec 2025). Reads 7,433 URLs from the master spreadsheet and
submits them to the Internet Archive's Save Page Now 2 (SPN2) API,
processing one Excel sheet at a time in order of URL count (smallest first).

Workflow:
  1. Load URLs from manuscripts-urls-wbm-archived.xlsx
  2. Process sheets: static_pages -> indexes -> search_literature ->
     search_extended -> show_text -> show_images_text -> show_manuscript
  3. For each URL: submit to SPN2 API, write WBM URL back to Excel
  4. Save Excel every 10 URLs and progress JSON after every URL

Resume capability:
  - Progress JSON saved after EVERY URL (archiving-progress.json)
  - Can resume from exact sheet + row after power outage or crash
  - Completed sheets are tracked so they're skipped on restart

Rate limiting:
  - 17s base delay between requests (safe for IA authenticated rate of 15 req/min)
  - Exponential backoff on transient failures (2x multiplier)
  - 5-minute pause + jitter on HTTP 429 rate limit responses

Requirements:
  - .env file with IA_ACCESS_KEY and IA_SECRET_KEY (in scripts/ folder)
  - openpyxl for Excel read/write
  - requests for HTTP

Usage:
  python SaveToWBM_manuscripts_bulk.py
"""

import os
import sys
import time
import json
import requests
import random
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    import openpyxl.utils
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# User-Agent
USER_AGENT = "KB-Archiver/1.0 (https://www.kb.nl/; manuscripts.kb.nl archiving project) Python/requests"

# Configuration
SCRIPT_DIR = Path(__file__).parent
ARTIFACTS_DIR = SCRIPT_DIR.parent
DATA_DIR = ARTIFACTS_DIR / "data" / "excelArchiving"
LOGS_DIR = ARTIFACTS_DIR / "logs"

# Input/Output files
EXCEL_FILE = ARTIFACTS_DIR.parent / "manuscripts-urls-wbm-archived.xlsx"
PROGRESS_FILE = DATA_DIR / "archiving-progress.json"
ERRORS_FILE = DATA_DIR / "archiving-errors.json"
LOG_FILE = LOGS_DIR / "excel-archiving-log.txt"

# Rate limiting - 17s is safe for authenticated users (15 req/min limit = 4s min)
BASE_DELAY = 17  # seconds between requests
MAX_RETRIES = 3
BACKOFF_MULTIPLIER = 2
RATE_LIMIT_PAUSE = 300  # 5 minutes pause on rate limit

# Load credentials from .env file
ENV_FILE = SCRIPT_DIR / ".env"

# Sheet processing order (smallest URL count first)
SHEET_ORDER = [
    "static_pages",       # 8 URLs
    "indexes",            # 9 URLs
    "search_literature",  # 397 URLs
    "search_extended",    # 806 URLs
    "show_text",          # 1,520 URLs
    "show_images_text",   # 2,322 URLs
    "show_manuscript",    # 2,371 URLs
]


def load_credentials():
    """Load IA S3 credentials from .env file."""
    if not ENV_FILE.exists():
        raise FileNotFoundError(f".env file not found at {ENV_FILE}")

    creds = {}
    with open(ENV_FILE, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                creds[key.strip()] = value.strip().strip("'\"")

    access_key = creds.get('IA_ACCESS_KEY')
    secret_key = creds.get('IA_SECRET_KEY')

    if not access_key or not secret_key:
        raise ValueError("IA_ACCESS_KEY and IA_SECRET_KEY must be set in .env file")

    return access_key, secret_key


def load_progress():
    """Load progress from previous run (for resume capability)."""
    if PROGRESS_FILE.exists():
        try:
            with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"  Warning: Could not load progress file: {e}")

    return {
        'start_time': None,
        'end_time': None,
        'duration_seconds': 0,
        'current_sheet': None,
        'current_row': 2,  # Start after header
        'sheets_completed': [],
        'total_urls': 0,
        'successful': 0,
        'failed': 0,
        'skipped': 0,
        'last_url': None,
        'last_job_id': None,
        'results': []
    }


def save_progress(progress, start_time=None):
    """Save progress for resume capability."""
    if start_time:
        progress['duration_seconds'] = (datetime.now() - start_time).total_seconds()
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, indent=2)


def load_errors():
    """Load error log from previous run."""
    if ERRORS_FILE.exists():
        try:
            with open(ERRORS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {'errors': []}


def save_errors(errors_data):
    """Save error log."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(ERRORS_FILE, 'w', encoding='utf-8') as f:
        json.dump(errors_data, f, indent=2)


def add_error(errors_data, url, sheet, row, error_msg, error_type="FAILED"):
    """Add error to error log."""
    errors_data['errors'].append({
        'url': url,
        'sheet': sheet,
        'row': row,
        'error': error_msg,
        'type': error_type,
        'timestamp': datetime.now().isoformat()
    })
    save_errors(errors_data)


def save_to_wayback(url, access_key, secret_key, retries=0):
    """Submit a single URL to the Wayback Machine via the SPN2 API.

    Returns a dict with 'success' (bool), and on success: 'job_id', 'wbm_url',
    'message'. On failure: 'error', and optionally 'rate_limited' (bool).
    """
    headers = {
        'Accept': 'application/json',
        'Authorization': f'LOW {access_key}:{secret_key}',
        'User-Agent': USER_AGENT
    }

    data = {
        'url': url,
    }

    try:
        response = requests.post(
            'https://web.archive.org/save',
            headers=headers,
            data=data,
            timeout=120
        )

        if response.status_code == 200:
            result = response.json()
            job_id = result.get('job_id', '')
            message = result.get('message', 'Submitted')

            # Construct WBM URL (will redirect to actual timestamp once indexed)
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            wbm_url = f"https://web.archive.org/web/{timestamp}/{url}"

            return {
                'success': True,
                'job_id': job_id,
                'wbm_url': wbm_url,
                'message': message,
                'url': url,
                'retries': retries
            }
        elif response.status_code == 429:
            return {
                'success': False,
                'error': 'HTTP 429 Rate Limited',
                'rate_limited': True,
                'url': url,
                'retries': retries
            }
        elif response.status_code == 523:
            # Origin unreachable - site may be down
            return {
                'success': False,
                'error': f'HTTP 523 Origin Unreachable (site may be down)',
                'url': url,
                'retries': retries
            }
        else:
            return {
                'success': False,
                'error': f'HTTP {response.status_code}',
                'response': response.text[:200] if response.text else '',
                'url': url,
                'retries': retries
            }
    except requests.exceptions.Timeout:
        return {
            'success': False,
            'error': 'Timeout (120s)',
            'url': url,
            'retries': retries
        }
    except requests.exceptions.ConnectionError as e:
        return {
            'success': False,
            'error': f'Connection Error: {str(e)[:100]}',
            'url': url,
            'retries': retries
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)[:200],
            'url': url,
            'retries': retries
        }


def setup_sheet_headers(ws):
    """Setup headers matching wiki-priority-urls-WBM.xlsx format."""
    headers = [
        ('URL', 60),
        ('Source', 22),
        ('WBM Status', 13),
        ('WBM Job ID', 20),
        ('WBM URL', 70),
        ('Archived At', 22),
        ('Error', 30)
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'


def update_excel_row(wb, sheet_name, row, wbm_url, status, job_id="", error=""):
    """Write SPN2 submission results back into a single Excel row.

    On first call for a sheet, restructures columns to match the standard
    schema (URL, Source, WBM Status, WBM Job ID, WBM URL, Archived At, Error)
    and colour-codes the status cell (green=OK, red=FAILED, yellow=SKIPPED).
    """
    ws = wb[sheet_name]

    # Check if headers need to be set up (column 2 should be 'Source')
    if ws.cell(row=1, column=2).value != 'Source':
        # Need to restructure - move URL data and set up headers
        # First, collect all URLs
        urls = []
        for r in range(2, ws.max_row + 1):
            url = ws.cell(row=r, column=1).value
            if url:
                urls.append(url)

        # Clear the sheet and set up proper headers
        setup_sheet_headers(ws)

        # Re-add URLs with Source column
        for r, url in enumerate(urls, 2):
            ws.cell(row=r, column=1, value=url)
            ws.cell(row=r, column=2, value=sheet_name)  # Source = sheet name

    # Update row with results
    # Column mapping: 1=URL, 2=Source, 3=WBM Status, 4=WBM Job ID, 5=WBM URL, 6=Archived At, 7=Error
    ws.cell(row=row, column=2, value=sheet_name)  # Source
    ws.cell(row=row, column=3, value=status)  # WBM Status
    ws.cell(row=row, column=4, value=job_id)  # WBM Job ID
    ws.cell(row=row, column=5, value=wbm_url)  # WBM URL
    ws.cell(row=row, column=6, value=datetime.now().isoformat())  # Archived At
    ws.cell(row=row, column=7, value=error if error else None)  # Error

    # Color code status
    status_cell = ws.cell(row=row, column=3)
    if status == 'OK':
        status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif status == 'FAILED':
        status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif status == 'SKIPPED':
        status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def get_sheet_url_count(wb, sheet_name):
    """Get number of URLs in a sheet."""
    if sheet_name not in wb.sheetnames:
        return 0
    ws = wb[sheet_name]
    return ws.max_row - 1  # Exclude header


def process_sheet(wb, sheet_name, access_key, secret_key, progress, errors_data, log_file, start_time):
    """Submit all URLs in one Excel sheet to the Wayback Machine.

    Iterates from the current row (for resume) to the last row, submitting
    each URL via save_to_wayback() with retry logic. Skips rows that already
    have a successful WBM URL. Updates the Excel, progress JSON, and error
    log after every URL. Returns True when the sheet is fully processed.
    """
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' not found, skipping...")
        return True  # Consider it complete

    ws = wb[sheet_name]
    total_rows = ws.max_row
    start_row = progress['current_row'] if progress['current_sheet'] == sheet_name else 2

    print(f"\n  Processing sheet: {sheet_name}")
    print(f"  Total URLs: {total_rows - 1}")
    print(f"  Starting from row: {start_row}")

    # Update progress
    progress['current_sheet'] = sheet_name

    for row in range(start_row, total_rows + 1):
        url = ws.cell(row=row, column=1).value

        if not url:
            progress['current_row'] = row + 1
            save_progress(progress)
            continue

        # Check if already processed (has WBM URL in column 5, status OK in column 3)
        existing_wbm = ws.cell(row=row, column=5).value
        existing_status = ws.cell(row=row, column=3).value

        if existing_wbm and existing_status == 'OK':
            print(f"    [{row}/{total_rows}] Already archived, skipping: ...{url[-40:]}")
            progress['skipped'] += 1
            progress['current_row'] = row + 1
            save_progress(progress, start_time)
            continue

        # Progress display
        sheet_progress = row - 1
        sheet_total = total_rows - 1
        pct = (sheet_progress * 100) // sheet_total if sheet_total > 0 else 0
        url_short = url[-50:] if len(url) > 50 else url
        print(f"    [{row}/{total_rows}] ({pct}%) ...{url_short}", end=" ", flush=True)

        # Submit to WBM with retry logic
        retry_count = 0
        current_delay = BASE_DELAY
        result = None

        while retry_count <= MAX_RETRIES:
            result = save_to_wayback(url, access_key, secret_key, retries=retry_count)

            if result['success']:
                break
            elif result.get('rate_limited'):
                if retry_count < MAX_RETRIES:
                    wait_time = RATE_LIMIT_PAUSE + random.uniform(30, 120)
                    print(f"\n    RATE LIMITED! Waiting {wait_time:.0f}s before retry...", end=" ", flush=True)
                    log_file.write(f"RATE LIMITED at {datetime.now().isoformat()} - waiting {wait_time:.0f}s\n")
                    time.sleep(wait_time)
                    retry_count += 1
                else:
                    break
            else:
                # Other errors - try once more with backoff
                if retry_count < MAX_RETRIES:
                    backoff = current_delay * BACKOFF_MULTIPLIER + random.uniform(5, 15)
                    print(f"RETRY ({retry_count + 1})...", end=" ", flush=True)
                    time.sleep(backoff)
                    retry_count += 1
                    current_delay = backoff
                else:
                    break

        # Process result
        result_entry = {
            'success': result['success'],
            'url': url,
            'sheet': sheet_name,
            'row': row,
            'timestamp': datetime.now().isoformat(),
            'retries': result.get('retries', 0)
        }

        if result['success']:
            job_id = result.get('job_id', 'no-job-id')
            wbm_url = result.get('wbm_url', '')
            message = result.get('message', '')
            print(f"OK | job: {job_id[-20:]} | {message[:50]}")
            update_excel_row(wb, sheet_name, row, wbm_url, 'OK', job_id)
            progress['successful'] += 1
            progress['last_url'] = url
            progress['last_job_id'] = job_id
            result_entry['job_id'] = job_id
            result_entry['wbm_url'] = wbm_url
            result_entry['message'] = message
            try:
                log_file.write(f"OK: {url}\n  Job: {job_id}\n  WBM: {wbm_url}\n  Msg: {message}\n")
                log_file.flush()
            except OSError:
                pass  # Log file issue, continue anyway
        else:
            error_msg = result.get('error', 'Unknown error')
            print(f"FAILED: {error_msg}")
            update_excel_row(wb, sheet_name, row, '', 'FAILED', '', error_msg)
            progress['failed'] += 1
            result_entry['error'] = error_msg
            add_error(errors_data, url, sheet_name, row, error_msg)
            try:
                log_file.write(f"FAILED: {url}\n  Error: {error_msg}\n")
                log_file.flush()
            except OSError:
                pass  # Log file issue, continue anyway

        # Add to results array
        progress['results'].append(result_entry)
        progress['current_row'] = row + 1

        # Save progress JSON after every URL (fast, for resume capability)
        save_progress(progress, start_time)

        # Save Excel every 10 successful URLs (streaming update)
        if progress['successful'] % 10 == 0:
            try:
                wb.save(EXCEL_FILE)
                print(f"    [Excel saved at {progress['successful']} archived]")
            except PermissionError:
                print(f"    WARNING: Could not save Excel (file open?)")

        print(f"    [Progress: {progress['successful']} OK, {progress['failed']} FAIL, {progress['skipped']} skip | {progress['duration_seconds']:.0f}s]")

        # Rate limiting delay
        if row < total_rows:
            delay = BASE_DELAY + random.uniform(2, 8)
            time.sleep(delay)

        # Periodic status update
        if (row - start_row + 1) % 25 == 0:
            print(f"\n    [Status: {progress['successful']} archived, {progress['failed']} failed, {progress['skipped']} skipped]")

    return True  # Sheet complete


def main():
    print("=" * 70)
    print("manuscripts.kb.nl Excel Sheet Archiver")
    print("Wayback Machine SPN2 API (Authenticated)")
    print("=" * 70)
    print()

    # Ensure directories exist
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)

    # Check Excel file exists
    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    print(f"Excel file: {EXCEL_FILE}")

    # Load credentials
    print("\nLoading credentials...")
    try:
        access_key, secret_key = load_credentials()
        print(f"  Access key: {access_key[:4]}...{access_key[-4:]}")
        print("  Credentials loaded successfully")
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    # Load Excel
    print("\nLoading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    print(f"  Sheets found: {wb.sheetnames}")

    # Show URL counts per sheet
    print("\n  URL counts per sheet:")
    sheet_counts = {}
    for sheet in SHEET_ORDER:
        count = get_sheet_url_count(wb, sheet)
        sheet_counts[sheet] = count
        print(f"    {sheet}: {count}")

    total_urls = sum(sheet_counts.values())
    print(f"\n  Total URLs to archive: {total_urls}")

    # Load progress
    progress = load_progress()
    errors_data = load_errors()

    # Handle migration from old format to new format
    if 'started_at' in progress and 'start_time' not in progress:
        # Migrate old format
        progress['start_time'] = progress.pop('started_at', None)
        progress['successful'] = progress.pop('total_archived', 0)
        progress['failed'] = progress.pop('total_failed', 0)
        progress['skipped'] = progress.pop('total_skipped', 0)
        if 'results' not in progress:
            progress['results'] = []
        if 'last_url' not in progress:
            progress['last_url'] = None
        if 'last_job_id' not in progress:
            progress['last_job_id'] = None

    if progress.get('start_time'):
        print(f"\n  Resuming from previous session:")
        print(f"    Started: {progress['start_time']}")
        print(f"    Current sheet: {progress.get('current_sheet')}")
        print(f"    Current row: {progress.get('current_row', 2)}")
        print(f"    Archived: {progress.get('successful', 0)}")
        print(f"    Failed: {progress.get('failed', 0)}")
        print(f"    Skipped: {progress.get('skipped', 0)}")
        start_time = datetime.fromisoformat(progress['start_time'])
    else:
        start_time = datetime.now()
        progress['start_time'] = start_time.isoformat()
        progress['total_urls'] = total_urls
        save_progress(progress, start_time)

    # Estimate time
    remaining_urls = total_urls - progress.get('successful', 0) - progress.get('failed', 0) - progress.get('skipped', 0)
    est_hours = (remaining_urls * BASE_DELAY) / 3600
    print(f"\n  Estimated time for {remaining_urls} remaining URLs: ~{est_hours:.1f} hours")
    print(f"  Delay between requests: {BASE_DELAY}s")

    print("\n" + "-" * 70)
    print("Starting archiving...")
    print("-" * 70)

    # Open log file
    log_mode = 'a' if progress['successful'] > 0 else 'w'
    with open(LOG_FILE, log_mode, encoding='utf-8') as log:
        if log_mode == 'w':
            log.write(f"Excel Archiving Log - manuscripts.kb.nl\n")
            log.write(f"Started: {datetime.now().isoformat()}\n")
            log.write(f"Excel file: {EXCEL_FILE}\n")
            log.write("=" * 70 + "\n\n")
        else:
            log.write(f"\n--- Resuming at {datetime.now().isoformat()} ---\n\n")

        # Process sheets in order
        for sheet_name in SHEET_ORDER:
            # Skip completed sheets
            if sheet_name in progress.get('sheets_completed', []):
                print(f"\n  Sheet '{sheet_name}' already completed, skipping...")
                continue

            # Process sheet
            completed = process_sheet(wb, sheet_name, access_key, secret_key, progress, errors_data, log, start_time)

            if completed:
                if 'sheets_completed' not in progress:
                    progress['sheets_completed'] = []
                progress['sheets_completed'].append(sheet_name)
                progress['current_row'] = 2  # Reset for next sheet
                save_progress(progress, start_time)

                # Save Excel at sheet completion
                try:
                    wb.save(EXCEL_FILE)
                    print(f"\n  [Excel saved - sheet complete]")
                except PermissionError:
                    print(f"\n  WARNING: Could not save Excel (file open?)")

                print(f"\n  Sheet '{sheet_name}' COMPLETE!")
                log.write(f"\n=== Sheet '{sheet_name}' completed at {datetime.now().isoformat()} ===\n\n")

        # Final summary
        end_time = datetime.now()

        log.write("\n" + "=" * 70 + "\n")
        log.write(f"ARCHIVING COMPLETE\n")
        log.write(f"Finished: {end_time.isoformat()}\n")
        log.write(f"Total archived: {progress['successful']}\n")
        log.write(f"Total failed: {progress['failed']}\n")
        log.write(f"Total skipped: {progress['skipped']}\n")

    # Final save
    wb.save(EXCEL_FILE)
    save_progress(progress)

    print("\n" + "=" * 70)
    print("ARCHIVING COMPLETE")
    print("=" * 70)
    print(f"Total archived: {progress['successful']}")
    print(f"Total failed: {progress['failed']}")
    print(f"Total skipped: {progress['skipped']}")
    print(f"\nExcel updated: {EXCEL_FILE}")
    print(f"Progress file: {PROGRESS_FILE}")
    print(f"Errors file: {ERRORS_FILE}")
    print(f"Log file: {LOG_FILE}")

    if progress['failed'] > 0:
        print(f"\nWARNING: {progress['failed']} URLs failed to archive!")
        print(f"Check {ERRORS_FILE} for details.")


if __name__ == "__main__":
    main()
