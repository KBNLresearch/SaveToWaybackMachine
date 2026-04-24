"""
Excel Writer for manuscripts.kb.nl URL Spider
Single sheet format (flat, no filters) with streaming saves every 25 URLs
Full resume capability after failure/connection loss
Created: 2025-12-10
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import (
    EXCEL_OUTPUT, EXCEL_HEADERS, CHECKPOINT_DIR, DATA_DIR,
    URL_CATEGORIES, PRIORITY_LEVELS
)

logger = logging.getLogger(__name__)

# State file for resume capability
STATE_FILE = DATA_DIR / "spider_state.json"


class URLRecord:
    """Represents a discovered URL with metadata."""

    def __init__(
        self,
        url: str,
        category: str = "OTHER",
        priority: str = "NORMAL",
        discovered: Optional[str] = None
    ):
        self.url = url
        self.path = urlparse(url).path
        self.category = category
        self.priority = priority
        self.discovered = discovered or datetime.now().isoformat()
        self.wbm_status = "pending"
        self.wbm_url = ""
        self.wbm_timestamp = ""

    def to_row(self) -> List[str]:
        """Convert to Excel row."""
        return [
            self.url,
            self.path,
            self.category,
            self.priority,
            self.discovered,
            self.wbm_status,
            self.wbm_url,
            self.wbm_timestamp,
        ]

    def to_dict(self) -> Dict:
        """Convert to dictionary for JSON serialization."""
        return {
            "url": self.url,
            "path": self.path,
            "category": self.category,
            "priority": self.priority,
            "discovered": self.discovered,
            "wbm_status": self.wbm_status,
            "wbm_url": self.wbm_url,
            "wbm_timestamp": self.wbm_timestamp,
        }

    @classmethod
    def from_dict(cls, data: Dict) -> "URLRecord":
        """Create from dictionary."""
        record = cls(
            url=data["url"],
            category=data.get("category", "OTHER"),
            priority=data.get("priority", "NORMAL"),
            discovered=data.get("discovered"),
        )
        record.wbm_status = data.get("wbm_status", "pending")
        record.wbm_url = data.get("wbm_url", "")
        record.wbm_timestamp = data.get("wbm_timestamp", "")
        return record


def classify_url(url: str) -> str:
    """Classify URL into functional category."""
    path = urlparse(url).path.lower()

    # Check static pages first (exact matches)
    if path in ["/", "/introduction", "/background", "/advanced"]:
        return "STATIC_PAGES"

    # Index pages
    if "/indexes/" in path:
        return "INDEX_PAGES"

    # Search results - manuscripts
    if "/search/manuscripts/" in path:
        return "SEARCH_RESULTS"

    # Search results - images
    if "/search/images_text/" in path or "/search/images/" in path:
        return "SEARCH_RESULTS"

    # Detail pages - manuscripts
    if "/show/manuscript/" in path or "/manuscript/" in path:
        return "MANUSCRIPT_DETAIL"

    # Detail pages - images
    if "/show/image/" in path or "/image/" in path:
        return "IMAGE_DETAIL"

    # Static assets
    if any(ext in path for ext in [".css", ".js", ".woff", ".svg"]):
        return "STATIC_ASSETS"

    return "OTHER"


def determine_priority(url: str, wiki_urls: Set[str], commons_urls: Set[str]) -> str:
    """Determine archiving priority based on Wikimedia presence."""
    if url in wiki_urls:
        return "WIKI_HIGH"
    elif url in commons_urls:
        return "COMMONS_HIGH"

    category = classify_url(url)
    if category == "MANUSCRIPT_DETAIL":
        return "MANUSCRIPT"
    elif category == "IMAGE_DETAIL":
        return "IMAGE"
    elif category == "INDEX_PAGES":
        return "INDEX"
    elif category == "SEARCH_RESULTS":
        return "SEARCH"
    elif category == "STATIC_PAGES":
        return "STATIC"

    return "NORMAL"


class ExcelWriter:
    """
    Streaming Excel writer with:
    - Save every 25 URLs
    - Full resume capability after failure
    - State file tracking for recovery
    """

    # Save to Excel every 25 URLs
    SAVE_INTERVAL = 25

    def __init__(
        self,
        output_path: Path = EXCEL_OUTPUT,
        resume: bool = True
    ):
        self.output_path = output_path
        self.resume = resume

        self.seen_urls: Set[str] = set()
        self.pending_records: List[URLRecord] = []  # Buffer before save
        self.total_written = 0
        self.urls_since_last_save = 0

        # Wiki priority sets (loaded later)
        self.wiki_urls: Set[str] = set()
        self.commons_urls: Set[str] = set()

        # Initialize or resume
        if resume and STATE_FILE.exists():
            self._load_state()
        else:
            self._init_fresh()

    def _init_fresh(self):
        """Initialize fresh workbook and state."""
        logger.info(f"Initializing fresh workbook: {self.output_path}")

        # Create new workbook
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "ALL_URLS"
        self._setup_sheet_headers(ws)
        self.wb.save(self.output_path)

        # Initialize state
        self._save_state()
        logger.info("Fresh workbook created")

    def _setup_sheet_headers(self, ws):
        """Set up headers and formatting for the sheet."""
        # Add headers
        for col, header in enumerate(EXCEL_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Set column widths
        column_widths = {
            "URL": 80,
            "Path": 50,
            "Category": 20,
            "Priority": 15,
            "Discovered": 25,
            "WBM_Status": 12,
            "WBM_URL": 80,
            "WBM_Timestamp": 18,
        }
        for col, header in enumerate(EXCEL_HEADERS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths.get(header, 15)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _load_state(self):
        """Load state from file and Excel for resume."""
        logger.info("Resuming from saved state...")

        # Load state file
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            state = json.load(f)

        self.total_written = state.get("total_written", 0)
        self.seen_urls = set(state.get("seen_urls", []))

        logger.info(f"Loaded state: {self.total_written} URLs written, {len(self.seen_urls)} seen")

        # Load workbook
        if self.output_path.exists():
            self.wb = load_workbook(self.output_path)
            # Double-check against Excel content
            self._verify_excel_state()
        else:
            # State file exists but Excel doesn't - recreate
            logger.warning("State file exists but Excel missing, recreating...")
            self._init_fresh()

    def _verify_excel_state(self):
        """Verify Excel matches state file, rebuild seen_urls if needed."""
        ws = self.wb["ALL_URLS"]
        excel_urls = set()

        for row in range(2, ws.max_row + 1):
            url = ws.cell(row=row, column=1).value
            if url:
                excel_urls.add(url)

        # If mismatch, trust Excel as source of truth
        if excel_urls != self.seen_urls:
            logger.warning(f"State mismatch: state has {len(self.seen_urls)}, Excel has {len(excel_urls)} URLs")
            self.seen_urls = excel_urls
            self.total_written = len(excel_urls)
            self._save_state()

        logger.info(f"Verified: {len(self.seen_urls)} URLs in Excel")

    def _save_state(self):
        """Save current state to JSON file for recovery."""
        state = {
            "timestamp": datetime.now().isoformat(),
            "total_written": self.total_written,
            "seen_urls": list(self.seen_urls),
            "excel_path": str(self.output_path),
        }

        # Atomic write with temp file
        temp_file = STATE_FILE.with_suffix(".tmp")
        with open(temp_file, "w", encoding="utf-8") as f:
            json.dump(state, f, indent=2, ensure_ascii=False)
        temp_file.replace(STATE_FILE)

        logger.debug(f"State saved: {self.total_written} URLs")

    def load_wiki_priority(self, wiki_file: Path, commons_file: Path):
        """Load Wikipedia and Commons priority URLs from JSON files."""
        if wiki_file.exists():
            with open(wiki_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                self.wiki_urls = set(data.get("urls", []))
            logger.info(f"Loaded {len(self.wiki_urls)} Wikipedia priority URLs")

        if commons_file.exists():
            with open(commons_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                self.commons_urls = set(data.get("urls", []))
            logger.info(f"Loaded {len(self.commons_urls)} Commons priority URLs")

    def add_url(self, url: str) -> bool:
        """
        Add a URL to be written.
        Returns True if new, False if duplicate.
        Automatically saves every 25 URLs.
        """
        if url in self.seen_urls:
            return False

        self.seen_urls.add(url)

        # Classify and prioritize
        category = classify_url(url)
        priority = determine_priority(url, self.wiki_urls, self.commons_urls)

        record = URLRecord(url=url, category=category, priority=priority)
        self.pending_records.append(record)
        self.urls_since_last_save += 1

        # Save every 25 URLs
        if self.urls_since_last_save >= self.SAVE_INTERVAL:
            self._flush_to_excel()

        return True

    def _flush_to_excel(self):
        """Write pending records to Excel and save."""
        if not self.pending_records:
            return

        logger.info(f"Saving {len(self.pending_records)} URLs to Excel...")

        # Write to Excel
        ws = self.wb["ALL_URLS"]
        start_row = ws.max_row + 1

        for i, record in enumerate(self.pending_records):
            row = start_row + i
            for col, value in enumerate(record.to_row(), start=1):
                ws.cell(row=row, column=col, value=value)

        self.total_written += len(self.pending_records)

        # Save Excel
        self.wb.save(self.output_path)

        # Clear buffer
        self.pending_records = []
        self.urls_since_last_save = 0

        # Save state
        self._save_state()

        logger.info(f"Saved. Total URLs in Excel: {self.total_written}")

    def force_save(self):
        """Force immediate save of any pending records."""
        if self.pending_records:
            self._flush_to_excel()

    def finalize(self):
        """Flush any remaining records and finalize."""
        self._flush_to_excel()
        self._save_state()
        logger.info(f"Finalized Excel with {self.total_written} URLs")

    def get_stats(self) -> Dict:
        """Get current statistics by reading from Excel."""
        category_counts = {}
        priority_counts = {}

        ws = self.wb["ALL_URLS"]
        for row in range(2, ws.max_row + 1):
            category = ws.cell(row=row, column=3).value
            priority = ws.cell(row=row, column=4).value
            if category:
                category_counts[category] = category_counts.get(category, 0) + 1
            if priority:
                priority_counts[priority] = priority_counts.get(priority, 0) + 1

        return {
            "total_urls": len(self.seen_urls),
            "total_written": self.total_written,
            "pending": len(self.pending_records),
            "by_category": category_counts,
            "by_priority": priority_counts,
        }


def create_empty_workbook():
    """Create an empty workbook with headers."""
    writer = ExcelWriter(resume=False)
    writer.finalize()
    return writer.output_path


if __name__ == "__main__":
    # Test: create empty workbook
    logging.basicConfig(level=logging.INFO)
    path = create_empty_workbook()
    print(f"Created empty workbook: {path}")
